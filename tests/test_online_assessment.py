from __future__ import annotations

import tempfile
import unittest
from datetime import timedelta
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook
from starlette.datastructures import UploadFile

from backend.main import (
    _assessment_provider_from_mcc_ticket,
    _build_assessment_provider_result,
    _style_assessment_workbook,
    _write_assessment_logic_sheet,
    _write_assessment_results,
    process_online_assessment_excels,
)


class OnlineAssessmentTests(unittest.TestCase):
    @staticmethod
    def _excel_upload(filename: str, data: pd.DataFrame) -> UploadFile:
        buffer = BytesIO()
        data.to_excel(buffer, index=False)
        buffer.seek(0)
        return UploadFile(buffer, filename=filename)

    @staticmethod
    def _csv_upload(filename: str, data: pd.DataFrame) -> UploadFile:
        buffer = BytesIO(data.to_csv(index=False).encode("utf-8-sig"))
        return UploadFile(buffer, filename=filename)

    def test_numeric_mcc_customer_uses_transfer_result_only(self) -> None:
        for value in (10086, 10086.0, "10086", "10086.0", "10086.00"):
            with self.subTest(value=value):
                self.assertEqual(
                    _assessment_provider_from_mcc_ticket(value, "转接-ivd服务1组-迪诺"),
                    "迪诺",
                )

        self.assertEqual(
            _assessment_provider_from_mcc_ticket(10010, "转接-ivd服务2组-科海"),
            "科海",
        )
        self.assertEqual(_assessment_provider_from_mcc_ticket(10001, "无需转接"), "")
        self.assertEqual(
            _assessment_provider_from_mcc_ticket("陈银亭", "转接-ivd服务2组-科海"),
            "迪诺",
        )
        self.assertEqual(
            _assessment_provider_from_mcc_ticket("未知客服", "转接-ivd服务1组-迪诺"),
            "",
        )

    def test_provider_result_applies_new_rules_and_weighted_ivd_total(self) -> None:
        mcc_call = pd.DataFrame(
            {
                "相关客服": ["迪诺"] * 10 + ["科海", "科海"],
                "响铃时间": [
                    0,
                    "0",
                    "0秒",
                    "00:00",
                    timedelta(0),
                    "10秒",
                    "15秒",
                    "20秒",
                    "",
                    "异常",
                    "10秒",
                    "20秒",
                ],
            }
        )
        video_service = pd.DataFrame(
            {
                "客服工号": ["59990080", "59990080", "s100003619"],
                "振铃时长": [10, 30, 20],
                "是否接通": ["是", "否", "是"],
                "失败原因": ["", "timeout", ""],
                "通话时间": [600, 300, 600],
                "评价": [5, 1, 5],
                "排队时长": [4, 0, 5],
            }
        )
        mcc_ticket = pd.DataFrame(
            {
                "受理客服": ["陈银亭", 10086, 10010.0, 10001, "未知客服"],
                "TS转接结果": [
                    "无需转接",
                    "转接-ivd服务1组-迪诺",
                    "转接-ivd服务2组-科海",
                    "无需转接",
                    "转接-ivd服务1组-迪诺",
                ],
                "创建时间": [
                    "2026-01-01 08:00:00",
                    "2026-01-01 08:00:00",
                    "2026-01-01 08:00:00",
                    "2026-01-01 08:00:00",
                    "2026-01-01 08:00:00",
                ],
                "关闭时间": [
                    "2026-01-01 09:00:00",
                    "2026-01-01 09:00:00",
                    "2026-01-01 10:00:00",
                    "2026-01-01 09:00:00",
                    "2026-01-01 09:00:00",
                ],
                "短信满意度结果": [1, 1, 3, 1, 1],
                "受理单(service call)状态": [
                    "技术支持解决",
                    "派工完成",
                    "技术支持解决",
                    "技术支持解决",
                    "技术支持解决",
                ],
            }
        )
        crm_video = pd.DataFrame(
            {
                "负责GTS编号": ["59990080", "59990080", "s100003619"],
                "负责GTS": ["", "", ""],
                "用户状态": ["技术支持解决", "转FSM平台派工", "技术支持解决"],
            }
        )
        quality = pd.DataFrame(
            {
                "代理商": ["迪诺", "迪诺", "科海"],
                "扣分项": ["", "1", ""],
            }
        )

        result = _build_assessment_provider_result(
            mcc_call, video_service, mcc_ticket, crm_video, quality
        ).set_index("分包商")

        self.assertEqual(list(result.index), ["迪诺", "科海", "IVD合计", "庆余堂", "尚肯"])

        self.assertEqual(result.loc["迪诺", "热线15秒分子"], 2)
        self.assertEqual(result.loc["迪诺", "热线15秒分母"], 5)
        self.assertEqual(result.loc["迪诺", "热线15秒接起率"], "40.00%")
        self.assertEqual(result.loc["IVD合计", "热线15秒分子"], 3)
        self.assertEqual(result.loc["IVD合计", "热线15秒分母"], 7)
        self.assertEqual(result.loc["IVD合计", "热线15秒接起率"], "42.86%")

        self.assertEqual(result.loc["迪诺", "在线处理热线工单数"], 2)
        self.assertEqual(result.loc["科海", "在线处理热线工单数"], 1)
        self.assertEqual(result.loc["IVD合计", "在线处理热线工单数"], 3)
        self.assertEqual(result.loc["IVD合计", "在线处理视频工单数"], 3)
        self.assertEqual(result.loc["IVD合计", "在线处理总工单数"], 6)
        self.assertEqual(result.loc["IVD合计", "平均在线工单处理时长"], "0.74小时")

        self.assertEqual(result.loc["迪诺", "热线满意度"], "100.00%")
        self.assertEqual(result.loc["科海", "热线满意度"], "60.00%")
        self.assertEqual(result.loc["IVD合计", "热线满意度"], "86.67%")
        self.assertEqual(result.loc["IVD合计", "视频满意度"], "73.33%")
        self.assertEqual(result.loc["IVD合计", "质检合格率"], "66.67%")
        self.assertEqual(result.loc["IVD合计", "热线解决率"], "66.67%")
        self.assertEqual(result.loc["IVD合计", "视频解决率"], "66.67%")
        self.assertEqual(result.loc["IVD合计", "视频占线率"], "66.67%")

        for _, row in result.iterrows():
            self.assertEqual(
                row["在线处理总工单数"],
                row["在线处理热线工单数"] + row["在线处理视频工单数"],
            )

        expected_count_columns = [
            "平均在线工单处理时长",
            "在线处理总工单数",
            "在线处理热线工单数",
            "在线处理视频工单数",
        ]
        actual_columns = list(result.columns)
        start = actual_columns.index("平均在线工单处理时长")
        self.assertEqual(actual_columns[start : start + 4], expected_count_columns)

    def test_export_contains_new_rows_columns_and_logic(self) -> None:
        provider = pd.DataFrame(
            [
                {
                    "分包商": name,
                    "平均在线工单处理时长": "N/A",
                    "在线处理总工单数": 0,
                    "在线处理热线工单数": 0,
                    "在线处理视频工单数": 0,
                }
                for name in ("迪诺", "科海", "IVD合计", "庆余堂", "尚肯")
            ]
        )
        branch = pd.DataFrame(
            [{"分公司": "测试分公司", "工单总数量": 1, "已取消工单数量": 0, "分公司工单取消率": "0.00%"}]
        )
        overall = pd.DataFrame(
            [{"指标": "测试", "分子说明": "", "分子值": 0, "分母说明": "", "分母值": 0, "结果": "N/A"}]
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            output = Path(temp_dir) / "assessment.xlsx"
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                _write_assessment_results(writer, branch, overall, provider)
                _write_assessment_logic_sheet(writer)
                _style_assessment_workbook(writer)

            workbook = load_workbook(output, read_only=True, data_only=True)
            try:
                self.assertEqual(workbook.sheetnames, ["计算结果", "计算逻辑"])
                result_values = list(workbook["计算结果"].values)
                flattened = [value for row in result_values for value in row if value is not None]
                self.assertIn("IVD合计", flattened)
                self.assertIn("在线处理总工单数", flattened)
                self.assertIn("在线处理热线工单数", flattened)
                self.assertIn("在线处理视频工单数", flattened)

                logic_text = "\n".join(
                    str(value)
                    for row in workbook["计算逻辑"].values
                    for value in row
                    if value is not None
                )
                self.assertIn("剔除响铃时间明确换算为0秒", logic_text)
                self.assertIn("TS转接结果", logic_text)
                self.assertIn("在线处理总工单数=热线工单数+视频工单数", logic_text)
                self.assertIn("不对两个分包商的百分比或平均时长做算术平均", logic_text)
            finally:
                workbook.close()

    def test_full_upload_pipeline_generates_consistent_preview_and_workbook(self) -> None:
        msp = pd.DataFrame(
            {
                "工单状态": ["处理中"],
                "分公司": ["测试分公司"],
                "申告渠道": ["热线申告"],
                "首次派单时间": ["2026-01-01 08:00:00"],
                "预约完成时间": ["2026-01-01 08:20:00"],
                "服务结束时间": ["2026-01-01 09:00:00"],
            }
        )
        mcc_call = pd.DataFrame(
            {
                "相关客服": ["迪诺", "迪诺", "科海"],
                "响铃时间": [0, 10, 20],
            }
        )
        video_service = pd.DataFrame(
            {
                "客服工号": ["59990080", "s100003619"],
                "振铃时长": [10, 20],
                "是否接通": ["是", "是"],
                "失败原因": ["", ""],
                "通话时间": [300, 600],
                "评价": [5, 4],
                "排队时长": [1, 4],
            }
        )
        mcc_ticket = pd.DataFrame(
            {
                "受理客服": [10086, 10010],
                "TS转接结果": ["转接-ivd服务1组-迪诺", "转接-ivd服务2组-科海"],
                "创建时间": ["2026-01-01 08:00:00", "2026-01-01 08:00:00"],
                "关闭时间": ["2026-01-01 09:00:00", "2026-01-01 10:00:00"],
                "短信满意度结果": [1, 2],
                "受理单(service call)状态": ["技术支持解决", "派工完成"],
            }
        )
        crm_video = pd.DataFrame(
            {
                "负责GTS编号": ["59990080", "s100003619"],
                "负责GTS": ["", ""],
                "用户状态": ["技术支持解决", "转FSM平台派工"],
            }
        )
        quality = pd.DataFrame(
            {
                "代理商": ["迪诺", "科海"],
                "扣分项": ["", "1"],
            }
        )

        uploads = [
            self._excel_upload("MSP工单总表.xlsx", msp),
            self._csv_upload("MCC通话记录.csv", mcc_call),
            self._excel_upload("视频服务记录.xlsx", video_service),
            self._excel_upload("MCC热线工单.xlsx", mcc_ticket),
            self._csv_upload("CRM视频工单.csv", crm_video),
            self._excel_upload("每日质检记录表.xlsx", quality),
        ]

        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                output_dir = Path(temp_dir)
                with patch("backend.main.OUTPUT_DIR", output_dir):
                    payload = process_online_assessment_excels(
                        *uploads,
                        include_source_sheets=True,
                    )

                self.assertEqual(payload["stats"]["mcc_ticket_total"], 2)
                self.assertEqual(
                    [row["分包商"] for row in payload["preview"]["rows"]],
                    ["迪诺", "科海", "IVD合计", "庆余堂", "尚肯"],
                )
                ivd_row = payload["preview"]["rows"][2]
                self.assertEqual(ivd_row["在线处理热线工单数"], 2)
                self.assertEqual(ivd_row["在线处理视频工单数"], 2)
                self.assertEqual(ivd_row["在线处理总工单数"], 4)

                output = output_dir / payload["filename"]
                self.assertTrue(output.is_file())
                workbook = load_workbook(output, read_only=True, data_only=True)
                try:
                    self.assertEqual(
                        workbook.sheetnames,
                        [
                            "源表-MSP工单总表",
                            "源表-MCC通话记录",
                            "源表-视频服务记录",
                            "源表-MCC热线工单",
                            "源表-CRM视频工单",
                            "源表-每日质检记录",
                            "计算结果",
                            "计算逻辑",
                        ],
                    )
                finally:
                    workbook.close()
        finally:
            for upload in uploads:
                upload.file.close()


if __name__ == "__main__":
    unittest.main()
