from __future__ import annotations

import os
import re
import socket
import uuid
import sys
import threading
import webbrowser
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from license_service import (
        FEATURES,
        LicenseError,
        activate_license,
        current_license,
        get_machine_code,
        is_license_required,
        license_to_info,
        require_valid_license,
    )
except ModuleNotFoundError:  # pragma: no cover - packaged import path
    from backend.license_service import (
        FEATURES,
        LicenseError,
        activate_license,
        current_license,
        get_machine_code,
        is_license_required,
        license_to_info,
        require_valid_license,
    )

try:
    from services.eclass.router import router as eclass_router
except ModuleNotFoundError:  # pragma: no cover - package import path used by launcher
    from backend.services.eclass.router import router as eclass_router


APP_NAME = "技术支持效率平台"


def _require_feature(feature: str | None = None) -> None:
    try:
        require_valid_license(feature)
    except LicenseError as exc:
        raise HTTPException(status_code=exc.status_code, detail=exc.message) from exc


def _bundle_root() -> Path:
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return Path(sys._MEIPASS)
    return Path(__file__).resolve().parent.parent


def _resolve_frontend_dist_dir() -> Path:
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        bundle_root = Path(sys._MEIPASS)
        candidates = [
            bundle_root / "frontend" / "dist",
            bundle_root / "dist",
            bundle_root,
        ]
        for candidate in candidates:
            if (candidate / "index.html").exists():
                return candidate

    source_candidate = Path(__file__).resolve().parent.parent / "frontend" / "dist"
    return source_candidate


def _resolve_output_dir() -> Path:
    if getattr(sys, "frozen", False):
        local_app_data = os.environ.get("LOCALAPPDATA")
        if local_app_data:
            return Path(local_app_data) / APP_NAME / "output"
        return Path.home() / "AppData" / "Local" / APP_NAME / "output"
    return Path(__file__).resolve().parent / "output"


def _find_free_port(start_port: int = 8000, attempts: int = 50) -> int:
    for port in range(start_port, start_port + attempts):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            try:
                sock.bind(("127.0.0.1", port))
            except OSError:
                continue
            return port
    raise RuntimeError("无法找到可用端口")


def _prepare_runtime_streams() -> None:
    if sys.stdout is not None and sys.stderr is not None:
        return

    log_dir = Path(os.environ.get("LOCALAPPDATA", Path.home())) / APP_NAME / "logs"
    log_dir.mkdir(parents=True, exist_ok=True)
    log_file = log_dir / "runtime.log"

    if sys.stdout is None:
        sys.stdout = open(log_file, "a", encoding="utf-8", buffering=1)
    if sys.stderr is None:
        sys.stderr = open(log_file, "a", encoding="utf-8", buffering=1)


FRONTEND_DIST_DIR = _resolve_frontend_dist_dir()
OUTPUT_DIR = _resolve_output_dir()
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

WORK_ORDER_SHEET_NAME = "工单报表 - 已筛选"
TARGET_PRODUCT_LINE = "IVD"
TARGET_STATUSES = {"服务执行中", "已派工", "已预约"}

WORK_ORDER_ALIASES = {
    "product_line": ["大产线", "产品线"],
    "status": ["工单状态", "状态"],
    "ticket_id": ["工单号", "服务工单号", "工单编号"],
}

QUALITY_ALIASES = {
    "ticket_id": ["前续工单", "工单号", "服务工单号", "前续工单号"],
}

ONLINE_BUSINESS_SHEETS = {
    "mcc": "MCC热线原始数据",
    "video": "视频工单原始数据",
    "msp": "MSP工单原始数据",
    "ivd_double": "双大",
    "ivd_ab": "县域+社办AB类",
    "ivd_jc": "JC+社办CD类",
}

ONLINE_REQUIRED_COLUMNS = {
    "mcc": ["客户ID", "客户名称", "受理单(service call)状态", "创建时间", "申告渠道", "VIP客户级别"],
    "video": ["创建日期", "用户状态", "客户代码", "客户", "产品线", "VIP客户级别"],
    "msp": ["创建时间", "申告渠道", "工作方式", "工单类型"],
    "ivd_double": ["客户代码", "客户名称"],
    "ivd_ab": ["客户代码", "客户名称"],
    "ivd_jc": ["客户数量"],
}

MCC_HOTLINE_CHANNELS = {"001_热线申告", "006_微信申告", "009_一键报修申告"}
MCC_TRANSFER_STATUS_FOR_ANNUAL_SHARE = {"技术支持解决", "派工完成", "取消", "无需技术支持"}
MCC_SOLVED_STATUS = "技术支持解决"
MCC_COVERAGE_STATUS = {"技术支持解决", "派工完成", "取消", "无需技术支持"}
VIDEO_TRANSFER_STATUS = {"技术支持解决", "无需技术支持", "转FSM平台派工", "座席解决"}
PROJECT_SOLVED_STATUS = {"无需技术支持", "座席解决", "坐席解决", "技术支持解决"}
HOTLINE_PROJECT_DENOMINATOR_STATUS = PROJECT_SOLVED_STATUS | {"派工完成"}
VIDEO_PROJECT_DENOMINATOR_STATUS = PROJECT_SOLVED_STATUS | {"转FSM平台派工"}
MCC_IVD_PRODUCT_LINE = "IVD"
VIDEO_IVD_PRODUCT_LINES = {"血球", "生化"}
MSP_REMOTE_CHANNEL = "远程申告"
MSP_CS_CHANNEL = "CS申告"
MSP_OEE_SOLVED_WORK_MODES = {"电话指导", "远程解决"}
MSP_REPAIR_TYPE = "维修服务"
MSP_PROBLEM_TYPE = "临床服务"

ASSESSMENT_ONLINE_CHANNELS = {"热线申告", "微信视频申告", "一键报修申告", "微信申告", "远程申告"}
ASSESSMENT_PROVIDERS = ["迪诺", "科海", "庆余堂", "尚肯"]
ASSESSMENT_PROVIDER_ALIASES = {
    "迪诺": ["迪诺", "四川迪诺", "陈银亭", "吴昌翼", "赵平", "马上", "杨洪"],
    "科海": ["科海", "河北科海", "温彦朝", "李勇健", "杨计正", "杨计正", "佟健阳", "佟建阳", "秦敬国"],
    "庆余堂": ["庆余堂", "龚世杰", "金苗苗"],
    "尚肯": ["尚肯", "西安尚肯", "孙欣"],
}
ASSESSMENT_STAFF_PROVIDER_MAP = {
    "59990080": "迪诺",
    "s100018415": "迪诺",
    "s100002594": "迪诺",
    "s100002595": "迪诺",
    "s100007767": "迪诺",
    "s100000290": "迪诺",
    "s100003619": "科海",
    "s100020775": "科海",
    "s100033043": "科海",
    "s100030770": "科海",
    "s100004649": "科海",
    "s100004229": "尚肯",
    "s100005735": "庆余堂",
    "s100003537": "庆余堂",
}


app = FastAPI(title=APP_NAME, version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.include_router(eclass_router)

if FRONTEND_DIST_DIR.exists():
    assets_dir = FRONTEND_DIST_DIR / "assets"
    if assets_dir.exists():
        app.mount("/assets", StaticFiles(directory=assets_dir), name="assets")


def _normalize_column_name(name: Any) -> str:
    return re.sub(r"\s+", "", str(name or "")).strip()


def _find_column(df: pd.DataFrame, aliases: list[str], label: str) -> str:
    normalized_to_original = {
        _normalize_column_name(column): column for column in df.columns
    }
    for alias in aliases:
        column = normalized_to_original.get(_normalize_column_name(alias))
        if column is not None:
            return column

    raise HTTPException(
        status_code=400,
        detail=f"未找到必需字段：{label}。当前表头：{', '.join(map(str, df.columns))}",
    )


def _normalize_cell_text(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def _normalize_ticket_id(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, (int, np.integer)):
        return str(int(value))
    if isinstance(value, (float, np.floating)) and float(value).is_integer():
        return str(int(value))

    text = str(value).strip()
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def _expand_ticket_ids(value: Any) -> set[str]:
    normalized = _normalize_ticket_id(value)
    if not normalized:
        return set()

    values = {normalized}
    for token in re.split(r"[,，;；、\s]+", normalized):
        token = _normalize_ticket_id(token)
        if token:
            values.add(token)
    return values


def _read_table(upload: UploadFile, label: str) -> pd.DataFrame:
    filename = upload.filename or ""
    lower_filename = filename.lower()
    if not lower_filename.endswith((".xlsx", ".xlsm", ".xls", ".csv")):
        raise HTTPException(status_code=400, detail=f"{label}必须是 Excel 或 CSV 文件")

    try:
        upload.file.seek(0)
        if lower_filename.endswith(".csv"):
            try:
                return pd.read_csv(upload.file, dtype=object, encoding="utf-8-sig")
            except UnicodeDecodeError:
                upload.file.seek(0)
                return pd.read_csv(upload.file, dtype=object, encoding="gbk")

        return pd.read_excel(upload.file, sheet_name=0, dtype=object)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"{label}解析失败：{exc}") from exc


def _read_assessment_table(upload: UploadFile, label: str, sheet_name: str | None = None) -> pd.DataFrame:
    filename = upload.filename or ""
    lower_filename = filename.lower()
    try:
        upload.file.seek(0)
        if lower_filename.endswith(".csv"):
            for encoding in ("utf-16", "utf-8-sig", "gb18030", "gbk"):
                for separator in ("\t", ","):
                    try:
                        upload.file.seek(0)
                        df = pd.read_csv(upload.file, dtype=object, encoding=encoding, sep=separator, index_col=False)
                        if len(df.columns) > 1:
                            return _normalize_dataframe(df)
                    except Exception:
                        continue
            raise ValueError("无法识别 CSV 编码或分隔符")

        if lower_filename.endswith((".xlsx", ".xlsm", ".xls")):
            excel = pd.ExcelFile(upload.file)
            target_sheet = sheet_name if sheet_name in excel.sheet_names else excel.sheet_names[0]
            return _normalize_dataframe(pd.read_excel(excel, sheet_name=target_sheet, dtype=object))

        raise ValueError("文件必须是 Excel 或 CSV")
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"{label}解析失败：{exc}") from exc


def _find_required_column_by_prefix(df: pd.DataFrame, prefix: str, label: str) -> str:
    for column in df.columns:
        if str(column).strip().startswith(prefix):
            return column
    raise HTTPException(
        status_code=400,
        detail=f"{label}缺少以“{prefix}”开头的字段。当前表头：{', '.join(map(str, df.columns))}",
    )


def _validate_xlsx_upload(upload: UploadFile, label: str, expected_filename: str) -> None:
    filename = upload.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail=f"{label}必须是 .xlsx 文件")
    if expected_filename not in filename:
        raise HTTPException(
            status_code=400,
            detail=f"{label}文件名应包含：{expected_filename}。当前文件名：{filename}",
        )


def _validate_xlsx_upload_any(
    upload: UploadFile, label: str, expected_filenames: list[str]
) -> None:
    filename = upload.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail=f"{label}必须是 .xlsx 文件")
    if not any(expected_filename in filename for expected_filename in expected_filenames):
        raise HTTPException(
            status_code=400,
            detail=f"{label}文件名应包含：{' 或 '.join(expected_filenames)}。当前文件名：{filename}",
        )


def _normalize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    normalized = df.copy()
    normalized.columns = [str(column).strip() for column in normalized.columns]
    for column in normalized.columns:
        if normalized[column].dtype == object:
            normalized[column] = normalized[column].map(
                lambda value: value.strip() if isinstance(value, str) else value
            )
    return normalized


def _read_required_excel_sheet(upload: UploadFile, label: str, sheet_name: str) -> pd.DataFrame:
    try:
        upload.file.seek(0)
        excel = pd.ExcelFile(upload.file)
        if sheet_name not in excel.sheet_names:
            raise HTTPException(
                status_code=400,
                detail=f"{label}缺少必需 Sheet：{sheet_name}。当前 Sheet：{', '.join(excel.sheet_names)}",
            )
        return _normalize_dataframe(pd.read_excel(excel, sheet_name=sheet_name, dtype=object))
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"{label}解析失败：{exc}") from exc


def _read_ivd_customer_sheets(upload: UploadFile) -> dict[str, pd.DataFrame]:
    try:
        upload.file.seek(0)
        excel = pd.ExcelFile(upload.file)
        required_sheets = [
            ONLINE_BUSINESS_SHEETS["ivd_double"],
            ONLINE_BUSINESS_SHEETS["ivd_ab"],
            ONLINE_BUSINESS_SHEETS["ivd_jc"],
        ]
        missing = [sheet for sheet in required_sheets if sheet not in excel.sheet_names]
        if missing:
            raise HTTPException(
                status_code=400,
                detail=f"IVD客户群数量文件缺少必需 Sheet：{', '.join(missing)}",
            )
        return {
            sheet: _normalize_dataframe(pd.read_excel(excel, sheet_name=sheet, dtype=object))
            for sheet in required_sheets
        }
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"IVD客户群数量文件解析失败：{exc}") from exc


def _ensure_columns(df: pd.DataFrame, required_columns: list[str], label: str) -> None:
    missing = [column for column in required_columns if column not in df.columns]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"{label}缺少必需字段：{', '.join(missing)}。当前表头：{', '.join(map(str, df.columns))}",
        )


def _normalize_identifier(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, (int, np.integer)):
        return str(int(value))
    if isinstance(value, (float, np.floating)) and float(value).is_integer():
        return str(int(value))
    text = str(value).strip()
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def _to_datetime_series(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series, errors="coerce")


def _filter_year(df: pd.DataFrame, date_col: str, year: int) -> pd.DataFrame:
    parsed_dates = _to_datetime_series(df[date_col])
    return df[parsed_dates.dt.year == year].copy()


def _filter_mcc_ivd_rows(df: pd.DataFrame) -> pd.DataFrame:
    product_line_columns = [
        column
        for column in df.columns
        if _normalize_column_name(column).startswith("大产线")
    ]
    if not product_line_columns:
        raise HTTPException(status_code=400, detail="MCC热线原始数据缺少大产线字段，无法计算 IVD 客户覆盖率")

    mask = pd.Series(False, index=df.index)
    for column in product_line_columns:
        mask = mask | (df[column].map(_normalize_cell_text) == MCC_IVD_PRODUCT_LINE)
    return df[mask].copy()


def _filter_video_ivd_rows(df: pd.DataFrame) -> pd.DataFrame:
    product_line = df["产品线"].map(_normalize_cell_text)
    return df[product_line.isin(VIDEO_IVD_PRODUCT_LINES)].copy()


def _detect_online_business_year(
    mcc_df: pd.DataFrame, video_df: pd.DataFrame, msp_df: pd.DataFrame
) -> int:
    years: set[int] = set()
    for df, date_col in ((mcc_df, "创建时间"), (video_df, "创建日期"), (msp_df, "创建时间")):
        parsed = _to_datetime_series(df[date_col])
        years.update(int(year) for year in parsed.dt.year.dropna().unique())
    if not years:
        raise HTTPException(status_code=400, detail="无法从创建时间字段识别有效年份")
    return max(years)


def _detect_online_business_years(
    mcc_df: pd.DataFrame, video_df: pd.DataFrame, msp_df: pd.DataFrame
) -> list[int]:
    years: set[int] = set()
    for df, date_col in ((mcc_df, "创建时间"), (video_df, "创建日期"), (msp_df, "创建时间")):
        parsed = _to_datetime_series(df[date_col])
        years.update(int(year) for year in parsed.dt.year.dropna().unique())
    if not years:
        raise HTTPException(status_code=400, detail="无法从创建时间字段识别有效年份")
    return sorted(years)


def _safe_divide(numerator: int | float, denominator: int | float) -> float | None:
    if not denominator:
        return None
    return float(numerator) / float(denominator)


def _format_ratio(value: float | None) -> str:
    if value is None:
        return "N/A"
    return f"{value:.6f}"


def _format_percent(value: float | None) -> str:
    if value is None:
        return "N/A"
    return f"{value * 100:.2f}%"


def _format_result_percent(numerator: int | float, denominator: int | float) -> str:
    return _format_percent(_safe_divide(numerator, denominator))


def _write_ivd_customer_source_sheets(
    writer: pd.ExcelWriter, ivd_sheets: dict[str, pd.DataFrame]
) -> None:
    for sheet_name, df in ivd_sheets.items():
        df.to_excel(writer, index=False, sheet_name=sheet_name)


def _build_customer_level_lookup(ivd_sheets: dict[str, pd.DataFrame]) -> tuple[set[str], set[str]]:
    double_ids = {
        _normalize_identifier(value)
        for value in ivd_sheets[ONLINE_BUSINESS_SHEETS["ivd_double"]]["客户代码"]
    }
    ab_ids = {
        _normalize_identifier(value)
        for value in ivd_sheets[ONLINE_BUSINESS_SHEETS["ivd_ab"]]["客户代码"]
    }
    return {value for value in double_ids if value}, {value for value in ab_ids if value}


def _get_customer_level(customer_id: Any, double_ids: set[str], ab_ids: set[str]) -> str:
    normalized = _normalize_identifier(customer_id)
    if normalized in double_ids:
        return "双大"
    if normalized in ab_ids:
        return "县域+社办AB类"
    return "JC+社办CD类"


def _build_mcc_customer_dedup_detail(
    mcc_year_df: pd.DataFrame, double_ids: set[str], ab_ids: set[str]
) -> pd.DataFrame:
    coverage_df = _filter_mcc_ivd_rows(mcc_year_df)
    coverage_df["__normalized_customer_id"] = coverage_df["客户ID"].map(_normalize_identifier)
    coverage_df = coverage_df[coverage_df["__normalized_customer_id"] != ""].copy()
    dedup_df = coverage_df.drop_duplicates("__normalized_customer_id", keep="first").copy()
    dedup_df.drop(columns=["__normalized_customer_id"], inplace=True)

    customer_id_index = dedup_df.columns.get_loc("客户ID") + 1
    customer_levels = dedup_df["客户ID"].map(lambda value: _get_customer_level(value, double_ids, ab_ids))
    if "客户层级" in dedup_df.columns:
        dedup_df["客户层级"] = customer_levels
    else:
        dedup_df.insert(customer_id_index, "客户层级", customer_levels)
    return dedup_df


def _make_metric_row(
    name: str,
    numerator_desc: str,
    numerator: int,
    denominator_desc: str,
    denominator: int,
    formula: str,
    remark: str = "",
) -> dict[str, Any]:
    ratio = _safe_divide(numerator, denominator)
    return {
        "指标名称": name,
        "分子说明": numerator_desc,
        "分子值": int(numerator),
        "分母说明": denominator_desc,
        "分母值": int(denominator),
        "计算公式": formula,
        "小数结果": _format_ratio(ratio),
        "百分比结果": _format_percent(ratio),
        "备注": remark if denominator else f"{remark}；分母为0".strip("；"),
    }


ONLINE_RESULT_COLUMNS = [
    "指标",
    "热线工单数量",
    "热线已解决工单数量",
    "视频工单数量",
    "视频已解决工单数量",
    "OEE工单总数量（转技术支持+MSP OEE数量）",
    "MSP OEE已解决工单数（电话指导+远程）",
    "CS申告（维修）",
    "CS申告（问题处理）",
    "CS申告（维修+问题处理）",
    "总客户数量",
    "双大客户数量",
    "县域+社办AB类客户数量",
    "（JC+社办CD）客户数量",
    "热线覆盖客户数量",
    "热线覆盖（双大）客户数量",
    "热线覆盖（县域+社办AB）客户数量",
    "热线覆盖（JC+社办）客户数量",
    "热线（整体客户）覆盖率",
    "热线（双大客户）覆盖率",
    "热线（县域+社办AB类客户）覆盖率",
    "热线（JC+社办）覆盖率",
    "热线工单占比",
    "OEE工单占比",
    "热线整体解决率",
    "热线解决率",
    "MSP OEE单整体解决率",
    "MSP OEE工单解决率",
]

VIDEO_RESULT_COLUMNS = [
    "指标",
    "热线工单数量",
    "热线已解决工单数量",
    "视频工单数量",
    "视频已解决工单数量",
    "OEE工单总数量（转技术支持+MSP OEE数量）",
    "MSP OEE已解决工单数（电话指导+远程）",
    "CS申告（维修）",
    "CS申告（问题处理）",
    "CS申告（维修+问题处理）",
    "总客户数量",
    "双大客户数量",
    "县域+社办AB类客户数量",
    "（JC+社办CD）客户数量",
    "视频覆盖客户数量",
    "视频覆盖（双大）客户数量",
    "视频覆盖（县域+社办AB）客户数量",
    "视频覆盖（JC+社办）客户数量",
    "视频（整体客户）覆盖率",
    "视频（双大客户）覆盖率",
    "视频（县域+社办AB类客户）覆盖率",
    "视频（JC+社办）覆盖率",
    "视频工单占比",
    "OEE工单占比",
    "视频整体解决率",
    "视频解决率",
]

COMBINED_COVERAGE_RESULT_COLUMNS = [
    "指标",
    "热线工单数量",
    "热线已解决工单数量",
    "视频工单数量",
    "视频已解决工单数量",
    "OEE工单总数量（转技术支持+MSP OEE数量）",
    "MSP OEE已解决工单数（电话指导+远程）",
    "CS申告（维修）",
    "CS申告（问题处理）",
    "CS申告（维修+问题处理）",
    "总客户数量",
    "双大客户数量",
    "县域+社办AB类客户数量",
    "（JC+社办CD）客户数量",
    "视频+热线覆盖客户数量",
    "视频+热线覆盖（双大）客户数量",
    "视频+热线覆盖（县域+社办AB）客户数量",
    "视频+热线覆盖（JC+社办）客户数量",
    "视频+热线（整体客户）覆盖率",
    "视频+热线（双大客户）覆盖率",
    "视频+热线（县域+社办AB类客户）覆盖率",
    "视频+热线（JC+社办）覆盖率",
    "整体在线工单占比",
    "在线解决率整体",
]


def _normalize_customer_name(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip()
    for separator in (" / ", "/", "（", "("):
        if separator in text:
            text = text.split(separator)[0].strip()
            break
    return re.sub(r"\s+", "", text)


def _build_customer_name_lookup(ivd_sheets: dict[str, pd.DataFrame]) -> tuple[set[str], set[str]]:
    double_names = {
        _normalize_customer_name(value)
        for value in ivd_sheets[ONLINE_BUSINESS_SHEETS["ivd_double"]]["客户名称"]
    }
    ab_names = {
        _normalize_customer_name(value)
        for value in ivd_sheets[ONLINE_BUSINESS_SHEETS["ivd_ab"]]["客户名称"]
    }
    return {value for value in double_names if value}, {value for value in ab_names if value}


def _make_customer_identity(customer_code: Any, customer_name: Any) -> str:
    normalized_code = _normalize_identifier(customer_code)
    if normalized_code:
        return normalized_code
    return f"NAME:{_normalize_customer_name(customer_name)}"


def _build_yearly_online_base_metrics(
    year: int,
    mcc_df: pd.DataFrame,
    video_df: pd.DataFrame,
    msp_df: pd.DataFrame,
) -> dict[str, int]:
    mcc_year_df = _filter_year(mcc_df, "创建时间", year)
    video_year_df = _filter_year(video_df, "创建日期", year)
    msp_year_df = _filter_year(msp_df, "创建时间", year)

    mcc_status = mcc_year_df["受理单(service call)状态"].map(_normalize_cell_text)
    video_status = video_year_df["用户状态"].map(_normalize_cell_text)
    msp_channel = msp_year_df["申告渠道"].map(_normalize_cell_text)
    msp_work_mode = msp_year_df["工作方式"].map(_normalize_cell_text)
    msp_type = msp_year_df["工单类型"].map(_normalize_cell_text)

    cs_repair_count = int(((msp_channel == MSP_CS_CHANNEL) & (msp_type == MSP_REPAIR_TYPE)).sum())
    cs_problem_count = int(((msp_channel == MSP_CS_CHANNEL) & (msp_type == MSP_PROBLEM_TYPE)).sum())
    hotline_project_solved_count = int(mcc_status.isin(PROJECT_SOLVED_STATUS).sum())
    hotline_project_denominator_count = int(mcc_status.isin(HOTLINE_PROJECT_DENOMINATOR_STATUS).sum())
    video_project_solved_count = int(video_status.isin(PROJECT_SOLVED_STATUS).sum())
    video_project_denominator_count = int(video_status.isin(VIDEO_PROJECT_DENOMINATOR_STATUS).sum())

    return {
        "hotline_ticket_count": int(len(mcc_year_df)),
        "hotline_solved_count": int((mcc_status == MCC_SOLVED_STATUS).sum()),
        "hotline_project_solved_count": hotline_project_solved_count,
        "hotline_project_denominator_count": hotline_project_denominator_count,
        "video_ticket_count": int(len(video_year_df)),
        "video_solved_count": int((video_status == MCC_SOLVED_STATUS).sum()),
        "video_project_solved_count": video_project_solved_count,
        "video_project_denominator_count": video_project_denominator_count,
        "msp_oee_count": int((msp_channel == MSP_REMOTE_CHANNEL).sum()),
        "msp_oee_solved_count": int(
            ((msp_channel == MSP_REMOTE_CHANNEL) & msp_work_mode.isin(MSP_OEE_SOLVED_WORK_MODES)).sum()
        ),
        "cs_repair_count": cs_repair_count,
        "cs_problem_count": cs_problem_count,
        "cs_total_count": cs_repair_count + cs_problem_count,
    }


def _build_video_coverage_counts(
    video_year_df: pd.DataFrame,
    double_ids: set[str],
    ab_ids: set[str],
    ab_names: set[str],
) -> dict[str, int]:
    coverage: dict[str, dict[str, bool]] = {}
    for _, row in video_year_df.iterrows():
        customer_code = _normalize_identifier(row["客户代码"])
        if not customer_code:
            continue
        identity = customer_code
        if identity not in coverage:
            coverage[identity] = {
                "double": False,
                "ab": False,
            }
        coverage[identity]["double"] = coverage[identity]["double"] or customer_code in double_ids
        coverage[identity]["ab"] = coverage[identity]["ab"] or customer_code in ab_ids

    total_count = len(coverage)
    double_count = sum(1 for item in coverage.values() if item["double"])
    ab_count = sum(1 for item in coverage.values() if item["ab"])
    return {
        "total": total_count,
        "double": double_count,
        "ab": ab_count,
        "jc": max(total_count - double_count - ab_count, 0),
    }


def _build_combined_video_hotline_coverage_counts(
    mcc_year_df: pd.DataFrame,
    video_year_df: pd.DataFrame,
    double_ids: set[str],
    ab_ids: set[str],
    ab_names: set[str],
) -> dict[str, int]:
    hotline_ids = {
        value
        for value in mcc_year_df["客户ID"].map(_normalize_identifier)
        if value
    }

    coverage = {customer_id: {"double": customer_id in double_ids} for customer_id in hotline_ids}
    for _, row in video_year_df.iterrows():
        customer_code = _normalize_identifier(row["客户代码"])
        if not customer_code:
            continue
        identity = customer_code

        coverage.setdefault(identity, {"double": False, "ab": False})
        coverage[identity]["double"] = coverage[identity]["double"] or customer_code in double_ids
        coverage[identity]["ab"] = coverage[identity].get("ab", False) or customer_code in ab_ids

    total_count = len(coverage)
    double_count = sum(1 for item in coverage.values() if item["double"])
    ab_count = sum(1 for identity in coverage if identity in ab_ids)

    return {
        "total": total_count,
        "double": double_count,
        "ab": ab_count,
        "jc": max(total_count - double_count - ab_count, 0),
    }


def _build_online_business_result_table(
    years: list[int],
    mcc_df: pd.DataFrame,
    video_df: pd.DataFrame,
    msp_df: pd.DataFrame,
    ivd_sheets: dict[str, pd.DataFrame],
    double_ids: set[str],
    ab_ids: set[str],
) -> pd.DataFrame:
    double_customer_count = int(len(ivd_sheets[ONLINE_BUSINESS_SHEETS["ivd_double"]]))
    ab_customer_count = int(len(ivd_sheets[ONLINE_BUSINESS_SHEETS["ivd_ab"]]))
    jc_customer_total = int(
        pd.to_numeric(
            ivd_sheets[ONLINE_BUSINESS_SHEETS["ivd_jc"]]["客户数量"], errors="coerce"
        ).fillna(0).sum()
    )
    ivd_customer_total = double_customer_count + ab_customer_count + jc_customer_total

    rows = []
    for year in years:
        mcc_year_df = _filter_year(mcc_df, "创建时间", year)
        mcc_ivd_year_df = _filter_mcc_ivd_rows(mcc_year_df)
        base_metrics = _build_yearly_online_base_metrics(year, mcc_df, video_df, msp_df)
        online_total_count = (
            base_metrics["hotline_ticket_count"]
            + base_metrics["video_ticket_count"]
            + base_metrics["msp_oee_count"]
            + base_metrics["cs_total_count"]
        )

        coverage_customer_ids = {
            value
            for value in mcc_ivd_year_df["客户ID"].map(_normalize_identifier)
            if value
        }
        coverage_customer_count = int(len(coverage_customer_ids))
        double_coverage_count = int(len(coverage_customer_ids & double_ids))
        ab_coverage_count = int(len(coverage_customer_ids & ab_ids))
        jc_coverage_count = int(max(coverage_customer_count - double_coverage_count - ab_coverage_count, 0))

        rows.append({
            "指标": f"{year}年",
            "热线工单数量": base_metrics["hotline_ticket_count"],
            "热线已解决工单数量": base_metrics["hotline_solved_count"],
            "视频工单数量": base_metrics["video_ticket_count"],
            "视频已解决工单数量": base_metrics["video_solved_count"],
            "OEE工单总数量（转技术支持+MSP OEE数量）": base_metrics["msp_oee_count"],
            "MSP OEE已解决工单数（电话指导+远程）": base_metrics["msp_oee_solved_count"],
            "CS申告（维修）": base_metrics["cs_repair_count"],
            "CS申告（问题处理）": base_metrics["cs_problem_count"],
            "CS申告（维修+问题处理）": base_metrics["cs_total_count"],
            "总客户数量": ivd_customer_total,
            "双大客户数量": double_customer_count,
            "县域+社办AB类客户数量": ab_customer_count,
            "（JC+社办CD）客户数量": jc_customer_total,
            "热线覆盖客户数量": coverage_customer_count,
            "热线覆盖（双大）客户数量": double_coverage_count,
            "热线覆盖（县域+社办AB）客户数量": ab_coverage_count,
            "热线覆盖（JC+社办）客户数量": jc_coverage_count,
            "热线（整体客户）覆盖率": _format_result_percent(coverage_customer_count, ivd_customer_total),
            "热线（双大客户）覆盖率": _format_result_percent(double_coverage_count, double_customer_count),
            "热线（县域+社办AB类客户）覆盖率": _format_result_percent(ab_coverage_count, ab_customer_count),
            "热线（JC+社办）覆盖率": _format_result_percent(jc_coverage_count, jc_customer_total),
            "热线工单占比": _format_result_percent(base_metrics["hotline_ticket_count"], online_total_count),
            "OEE工单占比": _format_result_percent(base_metrics["msp_oee_count"], online_total_count),
            "热线整体解决率": _format_result_percent(base_metrics["hotline_solved_count"], online_total_count),
            "热线解决率": _format_result_percent(
                base_metrics["hotline_project_solved_count"],
                base_metrics["hotline_project_denominator_count"],
            ),
            "MSP OEE单整体解决率": _format_result_percent(base_metrics["msp_oee_solved_count"], online_total_count),
            "MSP OEE工单解决率": _format_result_percent(base_metrics["msp_oee_solved_count"], base_metrics["msp_oee_count"]),
        })

    return pd.DataFrame(rows, columns=ONLINE_RESULT_COLUMNS)


def _build_video_business_result_table(
    years: list[int],
    mcc_df: pd.DataFrame,
    video_df: pd.DataFrame,
    msp_df: pd.DataFrame,
    ivd_sheets: dict[str, pd.DataFrame],
    double_ids: set[str],
    ab_ids: set[str],
    ab_names: set[str],
) -> pd.DataFrame:
    double_customer_count = int(len(ivd_sheets[ONLINE_BUSINESS_SHEETS["ivd_double"]]))
    ab_customer_count = int(len(ivd_sheets[ONLINE_BUSINESS_SHEETS["ivd_ab"]]))
    jc_customer_total = int(
        pd.to_numeric(
            ivd_sheets[ONLINE_BUSINESS_SHEETS["ivd_jc"]]["客户数量"], errors="coerce"
        ).fillna(0).sum()
    )
    ivd_customer_total = double_customer_count + ab_customer_count + jc_customer_total

    rows = []
    for year in years:
        base_metrics = _build_yearly_online_base_metrics(year, mcc_df, video_df, msp_df)
        video_year_df = _filter_video_ivd_rows(_filter_year(video_df, "创建日期", year))
        video_coverage = _build_video_coverage_counts(video_year_df, double_ids, ab_ids, ab_names)
        online_total_count = (
            base_metrics["hotline_ticket_count"]
            + base_metrics["video_ticket_count"]
            + base_metrics["msp_oee_count"]
            + base_metrics["cs_total_count"]
        )

        rows.append({
            "指标": f"{year}年",
            "热线工单数量": base_metrics["hotline_ticket_count"],
            "热线已解决工单数量": base_metrics["hotline_solved_count"],
            "视频工单数量": base_metrics["video_ticket_count"],
            "视频已解决工单数量": base_metrics["video_solved_count"],
            "OEE工单总数量（转技术支持+MSP OEE数量）": base_metrics["msp_oee_count"],
            "MSP OEE已解决工单数（电话指导+远程）": base_metrics["msp_oee_solved_count"],
            "CS申告（维修）": base_metrics["cs_repair_count"],
            "CS申告（问题处理）": base_metrics["cs_problem_count"],
            "CS申告（维修+问题处理）": base_metrics["cs_total_count"],
            "总客户数量": ivd_customer_total,
            "双大客户数量": double_customer_count,
            "县域+社办AB类客户数量": ab_customer_count,
            "（JC+社办CD）客户数量": jc_customer_total,
            "视频覆盖客户数量": video_coverage["total"],
            "视频覆盖（双大）客户数量": video_coverage["double"],
            "视频覆盖（县域+社办AB）客户数量": video_coverage["ab"],
            "视频覆盖（JC+社办）客户数量": video_coverage["jc"],
            "视频（整体客户）覆盖率": _format_result_percent(video_coverage["total"], ivd_customer_total),
            "视频（双大客户）覆盖率": _format_result_percent(video_coverage["double"], double_customer_count),
            "视频（县域+社办AB类客户）覆盖率": _format_result_percent(video_coverage["ab"], ab_customer_count),
            "视频（JC+社办）覆盖率": _format_result_percent(video_coverage["jc"], jc_customer_total),
            "视频工单占比": _format_result_percent(base_metrics["video_ticket_count"], online_total_count),
            "OEE工单占比": _format_result_percent(base_metrics["msp_oee_count"], online_total_count),
            "视频整体解决率": _format_result_percent(base_metrics["video_solved_count"], online_total_count),
            "视频解决率": _format_result_percent(
                base_metrics["video_project_solved_count"],
                base_metrics["video_project_denominator_count"],
            ),
        })

    return pd.DataFrame(rows, columns=VIDEO_RESULT_COLUMNS)


def _build_combined_coverage_result_table(
    years: list[int],
    mcc_df: pd.DataFrame,
    video_df: pd.DataFrame,
    msp_df: pd.DataFrame,
    ivd_sheets: dict[str, pd.DataFrame],
    double_ids: set[str],
    ab_ids: set[str],
    ab_names: set[str],
) -> pd.DataFrame:
    double_customer_count = int(len(ivd_sheets[ONLINE_BUSINESS_SHEETS["ivd_double"]]))
    ab_customer_count = int(len(ivd_sheets[ONLINE_BUSINESS_SHEETS["ivd_ab"]]))
    jc_customer_total = int(
        pd.to_numeric(
            ivd_sheets[ONLINE_BUSINESS_SHEETS["ivd_jc"]]["客户数量"], errors="coerce"
        ).fillna(0).sum()
    )
    ivd_customer_total = double_customer_count + ab_customer_count + jc_customer_total

    rows = []
    for year in years:
        base_metrics = _build_yearly_online_base_metrics(year, mcc_df, video_df, msp_df)
        mcc_year_df = _filter_mcc_ivd_rows(_filter_year(mcc_df, "创建时间", year))
        video_year_df = _filter_video_ivd_rows(_filter_year(video_df, "创建日期", year))
        online_total_count = (
            base_metrics["hotline_ticket_count"]
            + base_metrics["video_ticket_count"]
            + base_metrics["msp_oee_count"]
            + base_metrics["cs_total_count"]
        )
        online_solution_denominator = (
            base_metrics["hotline_ticket_count"]
            + base_metrics["video_ticket_count"]
            + base_metrics["msp_oee_count"]
        )
        online_solution_numerator = (
            base_metrics["hotline_solved_count"]
            + base_metrics["video_solved_count"]
            + base_metrics["msp_oee_solved_count"]
        )
        combined_coverage = _build_combined_video_hotline_coverage_counts(
            mcc_year_df=mcc_year_df,
            video_year_df=video_year_df,
            double_ids=double_ids,
            ab_ids=ab_ids,
            ab_names=ab_names,
        )

        rows.append({
            "指标": f"{year}年",
            "热线工单数量": base_metrics["hotline_ticket_count"],
            "热线已解决工单数量": base_metrics["hotline_solved_count"],
            "视频工单数量": base_metrics["video_ticket_count"],
            "视频已解决工单数量": base_metrics["video_solved_count"],
            "OEE工单总数量（转技术支持+MSP OEE数量）": base_metrics["msp_oee_count"],
            "MSP OEE已解决工单数（电话指导+远程）": base_metrics["msp_oee_solved_count"],
            "CS申告（维修）": base_metrics["cs_repair_count"],
            "CS申告（问题处理）": base_metrics["cs_problem_count"],
            "CS申告（维修+问题处理）": base_metrics["cs_total_count"],
            "总客户数量": ivd_customer_total,
            "双大客户数量": double_customer_count,
            "县域+社办AB类客户数量": ab_customer_count,
            "（JC+社办CD）客户数量": jc_customer_total,
            "视频+热线覆盖客户数量": combined_coverage["total"],
            "视频+热线覆盖（双大）客户数量": combined_coverage["double"],
            "视频+热线覆盖（县域+社办AB）客户数量": combined_coverage["ab"],
            "视频+热线覆盖（JC+社办）客户数量": combined_coverage["jc"],
            "视频+热线（整体客户）覆盖率": _format_result_percent(combined_coverage["total"], ivd_customer_total),
            "视频+热线（双大客户）覆盖率": _format_result_percent(combined_coverage["double"], double_customer_count),
            "视频+热线（县域+社办AB类客户）覆盖率": _format_result_percent(combined_coverage["ab"], ab_customer_count),
            "视频+热线（JC+社办）覆盖率": _format_result_percent(combined_coverage["jc"], jc_customer_total),
            "整体在线工单占比": _format_result_percent(
                online_solution_denominator,
                online_total_count,
            ),
            "在线解决率整体": _format_result_percent(
                online_solution_numerator,
                online_solution_denominator,
            ),
        })

    return pd.DataFrame(rows, columns=COMBINED_COVERAGE_RESULT_COLUMNS)


def _style_online_result_sheet(
    writer: pd.ExcelWriter,
    header_rows: list[int],
    title_rows: list[int],
    sheet_name: str = "计算结果",
) -> None:
    worksheet = writer.book[sheet_name]
    thin = Side(style="thin", color="333333")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="F4F7FB")
    orange_fill = PatternFill("solid", fgColor="F4B183")
    customer_fill = PatternFill("solid", fgColor="BFEFEA")
    coverage_fill = PatternFill("solid", fgColor="FFF2CC")
    metric_fill = PatternFill("solid", fgColor="FFF2CC")
    highlight_fill = PatternFill("solid", fgColor="FFF200")
    title_fill = PatternFill("solid", fgColor="EAF3FF")

    fill_by_column = {}
    for column_index in range(6, 8):
        fill_by_column[column_index] = orange_fill
    for column_index in range(11, 15):
        fill_by_column[column_index] = customer_fill
    for column_index in range(19, 23):
        fill_by_column[column_index] = coverage_fill
    for column_index in range(26, 30):
        fill_by_column[column_index] = highlight_fill
    for column_index in (23, 24, 25):
        fill_by_column[column_index] = metric_fill

    header_row_set = set(header_rows)
    title_row_set = set(title_rows)
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if cell.row in title_row_set:
                cell.font = Font(bold=True, size=16, color="111827")
                cell.fill = title_fill
            elif cell.row in header_row_set:
                cell.font = Font(bold=True, size=12, color="1F2937")
                cell.fill = fill_by_column.get(cell.column, header_fill)
            else:
                cell.font = Font(size=12, color="111827")
                cell.fill = fill_by_column.get(cell.column, PatternFill(fill_type=None))

    worksheet.freeze_panes = "B2"
    for header_row in header_rows:
        worksheet.row_dimensions[header_row].height = 78
    for title_row in title_rows:
        worksheet.row_dimensions[title_row].height = 30
    width_by_column = {
        1: 12,
        6: 18,
        7: 18,
        10: 16,
        11: 14,
        12: 14,
        13: 18,
        14: 18,
        15: 16,
        16: 18,
        17: 20,
        18: 18,
        19: 16,
        20: 17,
        21: 20,
        22: 17,
        23: 18,
        24: 18,
        26: 18,
        27: 18,
        28: 18,
        29: 18,
    }
    for column_index in range(1, worksheet.max_column + 1):
        width = width_by_column.get(column_index, 13)
        worksheet.column_dimensions[get_column_letter(column_index)].width = width


def _add_result_table_titles(
    writer: pd.ExcelWriter,
    table_specs: list[tuple[str, int, int]],
    sheet_name: str = "计算结果",
) -> None:
    worksheet = writer.book[sheet_name]
    for title, title_row, column_count in table_specs:
        worksheet.cell(row=title_row, column=1, value=title)
        worksheet.merge_cells(
            start_row=title_row,
            start_column=1,
            end_row=title_row,
            end_column=column_count,
        )


def _add_calculation_logic_sheet_legacy(writer: pd.ExcelWriter) -> None:
    workbook = writer.book
    if "计算逻辑" in workbook.sheetnames:
        del workbook["计算逻辑"]
    worksheet = workbook.create_sheet("计算逻辑")
    writer.sheets["计算逻辑"] = worksheet

    title_fill = PatternFill("solid", fgColor="0F766E")
    section_fill = PatternFill("solid", fgColor="DBEAFE")
    header_fill = PatternFill("solid", fgColor="EAF3FF")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    worksheet.merge_cells("A1:H1")
    worksheet["A1"] = "在线业务计算逻辑"
    worksheet["A1"].font = Font(bold=True, size=18, color="FFFFFF")
    worksheet["A1"].fill = title_fill
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 34

    worksheet.merge_cells("A2:H2")
    worksheet["A2"] = (
        "本 Sheet 记录当前程序实际使用的字段、筛选条件、去重规则、分子分母和 Excel 公式。"
        "用于审计“计算结果”Sheet 中热线数据、视频数据、热线+视频数据和图表区域的生成逻辑。"
    )
    worksheet["A2"].font = Font(size=11, color="334155")
    worksheet["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    worksheet.row_dimensions[2].height = 44

    headers = [
        "分类",
        "计算项",
        "数据来源",
        "年份/筛选条件",
        "分子或取值规则",
        "分母或比较基准",
        "去重/匹配规则",
        "输出/公式/备注",
    ]
    rows: list[list[str]] = []

    def section(name: str) -> None:
        rows.append([name, "", "", "", "", "", "", ""])

    def add(
        category: str,
        item: str,
        source: str,
        filter_rule: str,
        numerator_rule: str,
        denominator_rule: str,
        dedup_rule: str,
        output_rule: str,
    ) -> None:
        rows.append([
            category,
            item,
            source,
            filter_rule,
            numerator_rule,
            denominator_rule,
            dedup_rule,
            output_rule,
        ])

    section("一、输入校验与年度口径")
    add(
        "输入校验",
        "MCC热线原始数据",
        "上传文件：MCC热线原始数据.xlsx；Sheet：MCC热线原始数据",
        "文件必须为 .xlsx，文件名包含 MCC热线原始数据",
        "必需字段：客户ID、客户名称、受理单(service call)状态、创建时间、申告渠道、VIP客户级别",
        "无",
        "字段名按原始表头精确校验",
        "原样写入 Sheet：MCC热线原始数据",
    )
    add(
        "输入校验",
        "视频工单原始数据",
        "上传文件：视频工单原始数据.xlsx；Sheet：视频工单原始数据",
        "文件必须为 .xlsx，文件名包含 视频工单原始数据",
        "必需字段：创建日期、用户状态、客户代码、客户、VIP客户级别",
        "无",
        "字段名按原始表头精确校验",
        "原样写入 Sheet：视频工单原始数据",
    )
    add(
        "输入校验",
        "MSP工单原始数据",
        "上传文件：MSP工单原始数据.xlsx；Sheet：MSP工单原始数据",
        "文件必须为 .xlsx，文件名包含 MSP工单原始数据",
        "必需字段：创建时间、申告渠道、工作方式、工单类型",
        "无",
        "字段名按原始表头精确校验",
        "原样写入 Sheet：MSP工单原始数据",
    )
    add(
        "输入校验",
        "IVD客户群表",
        "上传文件：IVD客户群表.xlsx；兼容旧名 IVD客户群数量2025.xlsx",
        "必需 Sheet：双大、县域+社办AB类、JC+社办CD类",
        "双大/县域+社办AB类必需字段：客户代码、客户名称；JC+社办CD类必需字段：客户数量",
        "无",
        "双大、县域+社办AB类按客户代码匹配；AB 还会建立客户名称匹配集合",
        "按原始 Sheet 名分别写入：双大、县域+社办AB类、JC+社办CD类",
    )
    add(
        "年度口径",
        "可计算年份",
        "MCC.创建时间、视频.创建日期、MSP.创建时间",
        "解析三个日期字段中的年份并取并集，按年份升序输出",
        "无",
        "无",
        "无法识别年份时终止计算",
        "下载文件名使用最大年份：在线服务项目目标计算表_YYYY.xlsx",
    )

    section("二、基础状态、渠道和客户集合")
    add(
        "固定集合",
        "热线渠道集合",
        "MCC热线原始数据.申告渠道",
        "申告渠道 in [001_热线申告, 006_微信申告, 009_一键报修申告]",
        "用于热线渠道口径说明",
        "无",
        "当前统计表基础数量按 MCC 年度数据计算，未额外按渠道切分",
        "常量：MCC_HOTLINE_CHANNELS",
    )
    add(
        "固定集合",
        "MCC覆盖率状态集合",
        "MCC热线原始数据.受理单(service call)状态",
        "状态 in [技术支持解决, 派工完成, 取消, 无需技术支持]",
        "用于 MCC热线客户去重明细 Sheet",
        "无",
        "客户ID规范化后去空、去重",
        "常量：MCC_COVERAGE_STATUS",
    )
    add(
        "固定集合",
        "视频转技术支持状态集合",
        "视频工单原始数据.用户状态",
        "状态 in [技术支持解决, 无需技术支持, 转FSM平台派工, 座席解决]",
        "当前统计表的视频工单数量按年度视频全量行数计算",
        "无",
        "该集合保留为业务常量",
        "常量：VIDEO_TRANSFER_STATUS",
    )
    add(
        "客户集合",
        "双大客户",
        "IVD客户群表.双大",
        "读取客户代码、客户名称",
        "客户数量 = 双大 Sheet 行数",
        "无",
        "客户代码规范化后建立 double_ids；客户名称规范化后建立 double_names",
        "计算结果列：双大客户数量",
    )
    add(
        "客户集合",
        "县域+社办AB类客户",
        "IVD客户群表.县域+社办AB类",
        "读取客户代码、客户名称",
        "客户数量 = 县域+社办AB类 Sheet 行数",
        "无",
        "客户代码规范化后建立 ab_ids；客户名称规范化后建立 ab_names",
        "计算结果列：县域+社办AB类客户数量",
    )
    add(
        "客户集合",
        "JC+社办CD客户",
        "IVD客户群表.JC+社办CD类",
        "读取客户数量列",
        "客户数量 = 客户数量列转数字后求和",
        "无",
        "无",
        "计算结果列：（JC+社办CD）客户数量",
    )
    add(
        "客户集合",
        "总客户数量",
        "IVD客户群表",
        "无",
        "总客户数量 = 双大客户数量 + 县域+社办AB类客户数量 + JC+社办CD客户数量",
        "无",
        "无",
        "Excel公式：K行 = L行 + M行 + N行",
    )

    section("三、年度基础数量")
    add(
        "基础数量",
        "热线工单数量",
        "MCC热线原始数据",
        "创建时间年份 = 当前指标年份",
        "热线工单数量 = 年度 MCC 行数",
        "无",
        "不去重",
        "计算结果列：热线工单数量",
    )
    add(
        "基础数量",
        "热线已解决工单数量",
        "MCC热线原始数据",
        "创建时间年份 = 当前指标年份",
        "分子 = 受理单(service call)状态 == 技术支持解决 的年度 MCC 行数",
        "无",
        "不去重",
        "计算结果列：热线已解决工单数量",
    )
    add(
        "基础数量",
        "视频工单数量",
        "视频工单原始数据",
        "创建日期年份 = 当前指标年份",
        "视频工单数量 = 年度视频行数",
        "无",
        "不去重",
        "计算结果列：视频工单数量",
    )
    add(
        "基础数量",
        "视频已解决工单数量",
        "视频工单原始数据",
        "创建日期年份 = 当前指标年份",
        "分子 = 用户状态 == 技术支持解决 的年度视频行数",
        "无",
        "不去重",
        "计算结果列：视频已解决工单数量",
    )
    add(
        "基础数量",
        "OEE工单总数量",
        "MSP工单原始数据",
        "创建时间年份 = 当前指标年份",
        "数量 = 申告渠道 == 远程申告 的年度 MSP 行数",
        "无",
        "不去重",
        "计算结果列：OEE工单总数量（转技术支持+MSP OEE数量）",
    )
    add(
        "基础数量",
        "MSP OEE已解决工单数",
        "MSP工单原始数据",
        "创建时间年份 = 当前指标年份",
        "分子 = 申告渠道 == 远程申告 且 工作方式 in [电话指导, 远程解决] 的年度 MSP 行数",
        "无",
        "不去重",
        "计算结果列：MSP OEE已解决工单数（电话指导+远程）",
    )
    add(
        "基础数量",
        "CS申告（维修）",
        "MSP工单原始数据",
        "创建时间年份 = 当前指标年份",
        "数量 = 申告渠道 == CS申告 且 工单类型 == 维修服务 的年度 MSP 行数",
        "无",
        "不去重",
        "计算结果列：CS申告（维修）",
    )
    add(
        "基础数量",
        "CS申告（问题处理）",
        "MSP工单原始数据",
        "创建时间年份 = 当前指标年份",
        "数量 = 申告渠道 == CS申告 且 工单类型 == 临床服务 的年度 MSP 行数",
        "无",
        "不去重",
        "计算结果列：CS申告（问题处理）",
    )
    add(
        "基础数量",
        "CS申告（维修+问题处理）",
        "MSP工单原始数据",
        "创建时间年份 = 当前指标年份",
        "数量 = CS申告（维修） + CS申告（问题处理）",
        "无",
        "不去重",
        "Excel公式：J行 = H行 + I行",
    )
    add(
        "基础数量",
        "在线工单总量",
        "计算结果 Sheet 当前行",
        "当前指标年份",
        "在线工单总量 = 热线工单数量 + 视频工单数量 + OEE工单总数量 + CS申告（维修+问题处理）",
        "无",
        "不去重",
        "Excel公式片段：SUM(B行,D行,F行,J行)",
    )

    section("四、热线数据表覆盖与比例")
    add(
        "热线覆盖",
        "热线覆盖客户数量",
        "MCC热线原始数据",
        "创建时间年份 = 当前指标年份",
        "取年度 MCC 中非空客户ID",
        "总客户数量",
        "客户ID规范化后去重",
        "热线（整体客户）覆盖率：S行 = O行 / K行",
    )
    add(
        "热线覆盖",
        "热线覆盖（双大）客户数量",
        "MCC热线原始数据 + IVD客户群表.双大",
        "创建时间年份 = 当前指标年份",
        "年度 MCC 非空客户ID 与 double_ids 取交集",
        "双大客户数量",
        "客户ID规范化后去重",
        "热线（双大客户）覆盖率：T行 = P行 / L行",
    )
    add(
        "热线覆盖",
        "热线覆盖（县域+社办AB）客户数量",
        "MCC热线原始数据 + IVD客户群表.县域+社办AB类",
        "创建时间年份 = 当前指标年份",
        "年度 MCC 非空客户ID 与 ab_ids 取交集",
        "县域+社办AB类客户数量",
        "客户ID规范化后去重",
        "热线（县域+社办AB类客户）覆盖率：U行 = Q行 / M行",
    )
    add(
        "热线覆盖",
        "热线覆盖（JC+社办）客户数量",
        "MCC热线原始数据",
        "创建时间年份 = 当前指标年份",
        "热线覆盖（JC+社办）客户数量 = MAX(热线覆盖客户数量 - 双大覆盖 - AB覆盖, 0)",
        "JC+社办CD客户数量",
        "客户ID规范化后去重",
        "Excel公式：R行 = MAX(O行 - P行 - Q行, 0)；覆盖率：V行 = R行 / N行",
    )
    add(
        "热线比例",
        "热线工单占比",
        "计算结果.热线数据",
        "当前指标年份",
        "分子 = 热线工单数量",
        "分母 = 在线工单总量",
        "无",
        "Excel公式：W行 = B行 / SUM(B行,D行,F行,J行)",
    )
    add(
        "热线比例",
        "OEE工单占比",
        "计算结果.热线数据",
        "当前指标年份",
        "分子 = OEE工单总数量",
        "分母 = 在线工单总量",
        "无",
        "Excel公式：X行 = F行 / SUM(B行,D行,F行,J行)",
    )
    add(
        "热线比例",
        "热线整体解决率",
        "计算结果.热线数据",
        "当前指标年份",
        "分子 = 热线已解决工单数量",
        "分母 = 在线工单总量",
        "无",
        "Excel公式：Y行 = C行 / SUM(B行,D行,F行,J行)",
    )
    add(
        "热线比例",
        "热线解决率",
        "计算结果.热线数据",
        "当前指标年份",
        "分子 = 热线已解决工单数量",
        "分母 = 热线工单数量",
        "无",
        "Excel公式：Z行 = C行 / B行",
    )
    add(
        "热线比例",
        "MSP OEE单整体解决率",
        "计算结果.热线数据",
        "当前指标年份",
        "分子 = MSP OEE已解决工单数",
        "分母 = 在线工单总量",
        "无",
        "Excel公式：AA行 = G行 / SUM(B行,D行,F行,J行)",
    )
    add(
        "热线比例",
        "MSP OEE工单解决率",
        "计算结果.热线数据",
        "当前指标年份",
        "分子 = MSP OEE已解决工单数",
        "分母 = OEE工单总数量",
        "无",
        "Excel公式：AB行 = G行 / F行",
    )

    section("五、视频数据表覆盖与比例")
    add(
        "视频覆盖",
        "视频覆盖客户数量",
        "视频工单原始数据",
        "创建日期年份 = 当前指标年份",
        "对每行生成客户身份：优先客户代码；客户代码为空时使用规范化客户名称",
        "总客户数量",
        "按客户身份去重",
        "视频（整体客户）覆盖率：S行 = O行 / K行",
    )
    add(
        "视频覆盖",
        "视频覆盖（双大）客户数量",
        "视频工单原始数据 + IVD客户群表.双大",
        "创建日期年份 = 当前指标年份",
        "客户代码命中 double_ids",
        "双大客户数量",
        "按客户身份去重",
        "视频（双大客户）覆盖率：T行 = P行 / L行",
    )
    add(
        "视频覆盖",
        "视频覆盖（县域+社办AB）客户数量",
        "视频工单原始数据 + IVD客户群表.县域+社办AB类",
        "创建日期年份 = 当前指标年份",
        "客户代码命中 ab_ids，或客户名称命中 ab_names 且客户代码未命中 double_ids",
        "县域+社办AB类客户数量",
        "按客户身份去重",
        "视频（县域+社办AB类客户）覆盖率：U行 = Q行 / M行",
    )
    add(
        "视频覆盖",
        "视频覆盖（JC+社办）客户数量",
        "视频工单原始数据",
        "创建日期年份 = 当前指标年份",
        "视频覆盖（JC+社办）客户数量 = MAX(视频覆盖客户数量 - 双大覆盖 - AB覆盖, 0)",
        "JC+社办CD客户数量",
        "按客户身份去重",
        "Excel公式：R行 = MAX(O行 - P行 - Q行, 0)；覆盖率：V行 = R行 / N行",
    )
    add(
        "视频比例",
        "视频工单占比",
        "计算结果.视频数据",
        "当前指标年份",
        "分子 = 视频工单数量",
        "分母 = 在线工单总量",
        "无",
        "Excel公式：W行 = D行 / SUM(B行,D行,F行,J行)",
    )
    add(
        "视频比例",
        "OEE工单占比",
        "计算结果.视频数据",
        "当前指标年份",
        "分子 = OEE工单总数量",
        "分母 = 在线工单总量",
        "无",
        "Excel公式：X行 = F行 / SUM(B行,D行,F行,J行)",
    )
    add(
        "视频比例",
        "视频整体解决率",
        "计算结果.视频数据",
        "当前指标年份",
        "分子 = 视频已解决工单数量",
        "分母 = 在线工单总量",
        "无",
        "Excel公式：Y行 = E行 / SUM(B行,D行,F行,J行)",
    )
    add(
        "视频比例",
        "视频解决率",
        "计算结果.视频数据",
        "当前指标年份",
        "分子 = 视频已解决工单数量",
        "分母 = 视频工单数量",
        "无",
        "Excel公式：Z行 = E行 / D行",
    )

    section("六、热线+视频数据表覆盖")
    add(
        "热线+视频覆盖",
        "视频+热线覆盖客户数量",
        "MCC热线原始数据 + 视频工单原始数据",
        "创建时间/创建日期年份 = 当前指标年份",
        "热线客户身份集合 union 视频客户身份集合",
        "总客户数量",
        "热线用客户ID；视频优先客户代码，否则客户名称；按身份去重",
        "视频+热线（整体客户）覆盖率：S行 = O行 / K行",
    )
    add(
        "热线+视频覆盖",
        "视频+热线覆盖（双大）客户数量",
        "MCC热线原始数据 + 视频工单原始数据 + IVD客户群表.双大",
        "当前指标年份",
        "联合覆盖客户中客户代码命中 double_ids 的数量",
        "双大客户数量",
        "按身份去重",
        "视频+热线（双大客户）覆盖率：T行 = P行 / L行",
    )
    add(
        "热线+视频覆盖",
        "视频+热线覆盖（县域+社办AB）客户数量",
        "MCC热线原始数据 + 视频工单原始数据 + IVD客户群表.县域+社办AB类",
        "当前指标年份",
        "热线AB客户数 + 视频AB客户数 - 重复AB客户数",
        "县域+社办AB类客户数量",
        "重复AB客户：视频身份也在热线AB客户ID中，且视频 VIP客户级别为空",
        "视频+热线（县域+社办AB类客户）覆盖率：U行 = Q行 / M行",
    )
    add(
        "热线+视频覆盖",
        "视频+热线覆盖（JC+社办）客户数量",
        "MCC热线原始数据 + 视频工单原始数据",
        "当前指标年份",
        "视频+热线覆盖（JC+社办）客户数量 = MAX(联合覆盖客户数量 - 双大覆盖 - AB覆盖, 0)",
        "JC+社办CD客户数量",
        "按身份去重",
        "Excel公式：R行 = MAX(O行 - P行 - Q行, 0)；覆盖率：V行 = R行 / N行",
    )

    section("七、MCC热线客户去重明细")
    add(
        "去重明细",
        "MCC热线客户去重明细",
        "MCC热线原始数据",
        "仅使用最大年份数据；状态 in [技术支持解决, 派工完成, 取消, 无需技术支持]",
        "保留客户ID非空数据",
        "无",
        "按规范化客户ID去重，每个客户ID保留第一行",
        "在客户ID后插入客户层级列",
    )
    add(
        "去重明细",
        "客户层级",
        "MCC热线客户去重明细 + IVD客户群表",
        "无",
        "先匹配双大客户代码；未命中再匹配县域+社办AB类客户代码；都未命中则归为 JC+社办CD类",
        "无",
        "按客户ID规范化匹配",
        "可选值：双大、县域+社办AB类、JC+社办CD类",
    )

    section("八、图表和KPI区域")
    add(
        "KPI",
        "热线解决率",
        "计算结果.热线数据 最新年份",
        "取最新年份行",
        "直接引用热线解决率",
        "无",
        "无",
        "计算结果 Sheet KPI卡片：A12:E13",
    )
    add(
        "KPI",
        "视频解决率",
        "计算结果.视频数据 最新年份",
        "取最新年份行",
        "直接引用视频解决率",
        "无",
        "无",
        "计算结果 Sheet KPI卡片：F12:J13",
    )
    add(
        "KPI",
        "热线整体覆盖率",
        "计算结果.热线数据 最新年份",
        "取最新年份行",
        "直接引用热线（整体客户）覆盖率",
        "无",
        "无",
        "计算结果 Sheet KPI卡片：K12:O13",
    )
    add(
        "KPI",
        "总客户数",
        "计算结果.热线数据 最新年份",
        "取最新年份行",
        "直接引用总客户数量",
        "无",
        "无",
        "计算结果 Sheet KPI卡片：P12:T13",
    )
    add(
        "KPI",
        "总工单量",
        "计算结果 最新年份",
        "取最新年份行",
        "总工单量 = 热线工单数量 + 视频工单数量",
        "无",
        "不含 MSP OEE 和 CS 申告",
        "计算结果 Sheet KPI卡片：U12:Z13",
    )
    add(
        "图表",
        "工单来源结构分析",
        "隐藏 Sheet：图表数据",
        "取最新年份",
        "热线工单数量、视频工单数量",
        "总工单量 = 热线工单数量 + 视频工单数量",
        "无",
        "环形图，位置 A15，标题显示总工单量",
    )
    add(
        "图表",
        "热线与视频解决率对比",
        "隐藏 Sheet：图表数据",
        "取最新年份",
        "热线解决率、视频解决率",
        "无",
        "无",
        "柱状图，位置 J15",
    )
    add(
        "图表",
        "客户结构分布分析",
        "隐藏 Sheet：图表数据",
        "取最新年份",
        "JC+社办CD客户、县域+社办AB客户、双大客户",
        "无",
        "按客户数量从高到低排序",
        "横向条形图，位置 S15",
    )
    add(
        "图表",
        "热线客户覆盖率分析",
        "隐藏 Sheet：图表数据",
        "取最新年份",
        "热线整体覆盖率、热线双大覆盖率、热线县域+社办AB覆盖率、热线JC+社办CD覆盖率",
        "无",
        "无",
        "柱状图，位置 A29",
    )

    start_row = 4
    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(start_row, column_index, header)
        cell.font = Font(bold=True, size=11, color="0F172A")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_offset, row_values in enumerate(rows, start=start_row + 1):
        is_section = row_values[1:] == ["", "", "", "", "", "", ""]
        if is_section:
            worksheet.merge_cells(
                start_row=row_offset,
                start_column=1,
                end_row=row_offset,
                end_column=len(headers),
            )
            cell = worksheet.cell(row_offset, 1, row_values[0])
            cell.font = Font(bold=True, size=12, color="1E3A8A")
            cell.fill = section_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            for column_index in range(1, len(headers) + 1):
                worksheet.cell(row_offset, column_index).border = border
            worksheet.row_dimensions[row_offset].height = 24
            continue

        for column_index, value in enumerate(row_values, start=1):
            cell = worksheet.cell(row_offset, column_index, value)
            cell.font = Font(size=10, color="111827")
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        worksheet.row_dimensions[row_offset].height = 54

    worksheet.freeze_panes = "A5"
    worksheet.auto_filter.ref = f"A4:H{start_row + len(rows)}"
    column_widths = {
        1: 14,
        2: 24,
        3: 32,
        4: 36,
        5: 44,
        6: 34,
        7: 34,
        8: 44,
    }
    for column_index, width in column_widths.items():
        worksheet.column_dimensions[get_column_letter(column_index)].width = width


def _add_calculation_logic_sheet(writer: pd.ExcelWriter) -> None:
    workbook = writer.book
    if "计算逻辑" in workbook.sheetnames:
        del workbook["计算逻辑"]
    worksheet = workbook.create_sheet("计算逻辑")
    writer.sheets["计算逻辑"] = worksheet

    title_fill = PatternFill("solid", fgColor="0F766E")
    header_fill = PatternFill("solid", fgColor="EAF3FF")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    worksheet.merge_cells("A1:D1")
    worksheet["A1"] = "在线业务计算逻辑"
    worksheet["A1"].font = Font(bold=True, size=18, color="FFFFFF")
    worksheet["A1"].fill = title_fill
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 34

    worksheet.merge_cells("A2:D2")
    worksheet["A2"] = "本页根据程序实际计算步骤整理，说明“计算结果”中各指标的分子、分母、筛选和去重方式。"
    worksheet["A2"].font = Font(size=11, color="334155")
    worksheet["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    worksheet.row_dimensions[2].height = 38

    headers = ["序号", "指标", "计算逻辑", "表格来源"]
    rows = [
        [
            1,
            "在线解决率整体",
            "分子取热线中“技术支持解决”的工单数、视频中“技术支持解决”的工单数，以及 MSP 远程申告中工作方式为“电话指导”或“远程解决”的工单数之和；分母取热线工单数、视频工单数、MSP 远程申告工单数之和。",
            "MCC热线原始数据；视频工单原始数据；MSP工单原始数据",
        ],
        [
            2,
            "热线解决率（项目汇报）",
            "在 MCC 年度数据中按受理单状态统计：分子为“无需技术支持”“座席解决/坐席解决”“技术支持解决”的数量之和；分母为前述三类再加“派工完成”的数量之和。",
            "MCC热线原始数据",
        ],
        [
            3,
            "视频解决率（项目汇报）",
            "在视频年度数据中按用户状态统计：分子为“无需技术支持”“座席解决/坐席解决”“技术支持解决”的数量之和；分母为前述三类再加“转FSM平台派工”的数量之和。",
            "视频工单原始数据",
        ],
        [
            4,
            "OEE解决率",
            "分子为 MSP 年度远程申告中工作方式为“电话指导”或“远程解决”的工单数；分母为 MSP 年度远程申告工单数。",
            "MSP工单原始数据",
        ],
        [
            5,
            "热线工单占比",
            "分子为 MCC 年度工单总数；分母为热线工单数、视频工单数、MSP 远程申告数、MSP CS申告数之和。",
            "MCC热线原始数据；视频工单原始数据；MSP工单原始数据",
        ],
        [
            6,
            "视频工单占比",
            "分子为视频年度工单总数；分母为热线工单数、视频工单数、MSP 远程申告数、MSP CS申告数之和。",
            "MCC热线原始数据；视频工单原始数据；MSP工单原始数据",
        ],
        [
            7,
            "OEE工单占比",
            "分子为 MSP 年度远程申告工单数；分母为热线工单数、视频工单数、MSP 远程申告数、MSP CS申告数之和。",
            "MCC热线原始数据；视频工单原始数据；MSP工单原始数据",
        ],
        [
            8,
            "整体在线工单占比",
            "分子为热线工单数、视频工单数、MSP 远程申告数之和；分母为热线工单数、视频工单数、MSP 远程申告数、MSP CS申告数之和。",
            "MCC热线原始数据；视频工单原始数据；MSP工单原始数据",
        ],
        [
            9,
            "IVD整体客户覆盖率",
            "分子为 MCC 中大产线为 IVD 的客户ID与视频中产品线为“血球”或“生化”的客户代码合并后去重的数量；分母为双大、县域+社办AB类、JC+社办CD类客户数量之和。",
            "IVD客户群表；MCC热线原始数据；视频工单原始数据",
        ],
        [
            10,
            "双大客户覆盖率",
            "分子为覆盖客户中命中 IVD客户群表“双大”客户代码的数量；分母为“双大”Sheet 的客户行数。",
            "IVD客户群表；MCC热线原始数据；视频工单原始数据",
        ],
        [
            11,
            "县域+社办AB类客户覆盖率",
            "分子为覆盖客户中命中 IVD客户群表“县域+社办AB类”客户代码的数量；分母为该 Sheet 的客户行数。",
            "IVD客户群表；MCC热线原始数据；视频工单原始数据",
        ],
        [
            12,
            "JC+社办覆盖率",
            "分子为覆盖客户总数减去双大覆盖数和县域+社办AB覆盖数；分母为“JC+社办CD类”Sheet 中“客户数量”列求和。",
            "IVD客户群表；MCC热线原始数据；视频工单原始数据",
        ],
        [
            13,
            "IVD整体客户覆盖率（热线）",
            "分子只取 MCC 年度数据中大产线为 IVD 且客户ID非空的去重客户数；分母为 IVD 全部客户数量。",
            "IVD客户群表；MCC热线原始数据",
        ],
        [
            14,
            "双大客户覆盖率（热线）",
            "分子为热线覆盖客户中命中“双大”客户代码的数量；分母为“双大”Sheet 的客户行数。",
            "IVD客户群表；MCC热线原始数据",
        ],
        [
            15,
            "县域+社办AB类客户覆盖率（热线）",
            "分子为热线覆盖客户中命中“县域+社办AB类”客户代码的数量；分母为该 Sheet 的客户行数。",
            "IVD客户群表；MCC热线原始数据",
        ],
        [
            16,
            "JC+社办覆盖率（热线）",
            "分子为热线覆盖客户总数减去热线双大覆盖数和热线县域+社办AB覆盖数；分母为“JC+社办CD类”客户数量。",
            "IVD客户群表；MCC热线原始数据",
        ],
        [
            17,
            "IVD整体客户覆盖率（视频）",
            "分子只取视频年度数据中产品线为“血球”或“生化”且客户代码非空的去重客户数；分母为 IVD 全部客户数量。",
            "IVD客户群表；视频工单原始数据",
        ],
        [
            18,
            "双大客户覆盖率（视频）",
            "分子为视频覆盖客户中命中“双大”客户代码的数量；分母为“双大”Sheet 的客户行数。",
            "IVD客户群表；视频工单原始数据",
        ],
        [
            19,
            "县域+社办AB类客户覆盖率（视频）",
            "分子为视频覆盖客户中命中“县域+社办AB类”客户代码的数量；分母为该 Sheet 的客户行数。",
            "IVD客户群表；视频工单原始数据",
        ],
        [
            20,
            "JC+社办覆盖率（视频）",
            "分子为视频覆盖客户总数减去视频双大覆盖数和视频县域+社办AB覆盖数；分母为“JC+社办CD类”客户数量。",
            "IVD客户群表；视频工单原始数据",
        ],
    ]

    start_row = 4
    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(start_row, column_index, header)
        cell.font = Font(bold=True, size=11, color="0F172A")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_offset, row_values in enumerate(rows, start=start_row + 1):
        for column_index, value in enumerate(row_values, start=1):
            cell = worksheet.cell(row_offset, column_index, value)
            cell.font = Font(size=10, color="111827")
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        worksheet.row_dimensions[row_offset].height = 62

    worksheet.freeze_panes = "A5"
    worksheet.auto_filter.ref = f"A4:D{start_row + len(rows)}"
    for column_index, width in {1: 10, 2: 28, 3: 82, 4: 42}.items():
        worksheet.column_dimensions[get_column_letter(column_index)].width = width


def _apply_online_result_formulas(
    writer: pd.ExcelWriter,
    header_row: int,
    row_count: int,
    sheet_name: str = "计算结果",
) -> None:
    worksheet = writer.book[sheet_name]
    for row_index in range(header_row + 1, header_row + row_count + 1):
        total_online = f"SUM(B{row_index},D{row_index},F{row_index},J{row_index})"
        formula_by_column = {
            "J": f"=H{row_index}+I{row_index}",
            "K": f"=L{row_index}+M{row_index}+N{row_index}",
            "R": f"=MAX(O{row_index}-P{row_index}-Q{row_index},0)",
            "S": f'=IFERROR(O{row_index}/K{row_index},"N/A")',
            "T": f'=IFERROR(P{row_index}/L{row_index},"N/A")',
            "U": f'=IFERROR(Q{row_index}/M{row_index},"N/A")',
            "V": f'=IFERROR(R{row_index}/N{row_index},"N/A")',
            "W": f'=IFERROR(B{row_index}/{total_online},"N/A")',
            "X": f'=IFERROR(F{row_index}/{total_online},"N/A")',
            "Y": f'=IFERROR(C{row_index}/{total_online},"N/A")',
            "AA": f'=IFERROR(G{row_index}/{total_online},"N/A")',
            "AB": f'=IFERROR(G{row_index}/F{row_index},"N/A")',
        }
        for column, formula in formula_by_column.items():
            worksheet[f"{column}{row_index}"] = formula

        for column in ("S", "T", "U", "V", "W", "X", "Y", "AA", "AB"):
            worksheet[f"{column}{row_index}"].number_format = "0.00%"


def _apply_video_result_formulas(
    writer: pd.ExcelWriter,
    header_row: int,
    row_count: int,
    sheet_name: str = "计算结果",
) -> None:
    worksheet = writer.book[sheet_name]
    for row_index in range(header_row + 1, header_row + row_count + 1):
        total_online = f"SUM(B{row_index},D{row_index},F{row_index},J{row_index})"
        formula_by_column = {
            "J": f"=H{row_index}+I{row_index}",
            "K": f"=L{row_index}+M{row_index}+N{row_index}",
            "R": f"=MAX(O{row_index}-P{row_index}-Q{row_index},0)",
            "S": f'=IFERROR(O{row_index}/K{row_index},"N/A")',
            "T": f'=IFERROR(P{row_index}/L{row_index},"N/A")',
            "U": f'=IFERROR(Q{row_index}/M{row_index},"N/A")',
            "V": f'=IFERROR(R{row_index}/N{row_index},"N/A")',
            "W": f'=IFERROR(D{row_index}/{total_online},"N/A")',
            "X": f'=IFERROR(F{row_index}/{total_online},"N/A")',
            "Y": f'=IFERROR(E{row_index}/{total_online},"N/A")',
        }
        for column, formula in formula_by_column.items():
            worksheet[f"{column}{row_index}"] = formula

        for column in ("S", "T", "U", "V", "W", "X", "Y"):
            worksheet[f"{column}{row_index}"].number_format = "0.00%"


def _apply_combined_coverage_result_formulas(
    writer: pd.ExcelWriter,
    header_row: int,
    row_count: int,
    sheet_name: str = "计算结果",
) -> None:
    worksheet = writer.book[sheet_name]
    for row_index in range(header_row + 1, header_row + row_count + 1):
        total_online = f"SUM(B{row_index},D{row_index},F{row_index},J{row_index})"
        online_solution_denominator = f"SUM(B{row_index},D{row_index},F{row_index})"
        formula_by_column = {
            "J": f"=H{row_index}+I{row_index}",
            "K": f"=L{row_index}+M{row_index}+N{row_index}",
            "R": f"=MAX(O{row_index}-P{row_index}-Q{row_index},0)",
            "S": f'=IFERROR(O{row_index}/K{row_index},"N/A")',
            "T": f'=IFERROR(P{row_index}/L{row_index},"N/A")',
            "U": f'=IFERROR(Q{row_index}/M{row_index},"N/A")',
            "V": f'=IFERROR(R{row_index}/N{row_index},"N/A")',
            "W": f'=IFERROR({online_solution_denominator}/{total_online},"N/A")',
            "X": f'=IFERROR(SUM(C{row_index},E{row_index},G{row_index})/{online_solution_denominator},"N/A")',
        }
        for column, formula in formula_by_column.items():
            worksheet[f"{column}{row_index}"] = formula

        for column in ("S", "T", "U", "V", "W", "X"):
            worksheet[f"{column}{row_index}"].number_format = "0.00%"


def _percent_to_float(value: Any) -> float:
    if value is None or pd.isna(value):
        return 0.0
    if isinstance(value, str):
        text = value.strip()
        if text.endswith("%"):
            try:
                return float(text[:-1]) / 100
            except ValueError:
                return 0.0
        try:
            return float(text)
        except ValueError:
            return 0.0
    return float(value)


def _style_dashboard_card(worksheet, cell_range: str, title: str, value: str, fill_color: str) -> None:
    thin = Side(style="thin", color="4B5563")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    start_cell = cell_range.split(":")[0]
    worksheet.merge_cells(cell_range)
    cell = worksheet[start_cell]
    cell.value = f"{title}\n{value}"
    cell.font = Font(bold=True, size=14, color="F8FAFC")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.fill = PatternFill("solid", fgColor=fill_color)
    for row in worksheet[cell_range]:
        for range_cell in row:
            range_cell.border = border


def _add_chart_section_label(
    worksheet,
    cell_range: str,
    title: str,
    description: str,
) -> None:
    start_cell = cell_range.split(":")[0]
    worksheet.merge_cells(cell_range)
    cell = worksheet[start_cell]
    cell.value = f"{title}\n{description}"
    cell.font = Font(bold=True, size=13, color="1F2937")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.fill = PatternFill("solid", fgColor="EAF3FF")


def _write_dashboard_data_sheet(
    writer: pd.ExcelWriter,
    latest_result: dict[str, Any],
    latest_video_result: dict[str, Any],
) -> None:
    workbook = writer.book
    if "图表数据" in workbook.sheetnames:
        del workbook["图表数据"]
    sheet = workbook.create_sheet("图表数据")
    sheet.sheet_state = "hidden"

    hotline_count = int(latest_result["热线工单数量"])
    video_count = int(latest_video_result["视频工单数量"])
    customer_structure = [
        ("JC+社办CD客户", int(latest_result["（JC+社办CD）客户数量"])),
        ("县域+社办AB客户", int(latest_result["县域+社办AB类客户数量"])),
        ("双大客户", int(latest_result["双大客户数量"])),
    ]
    customer_structure.sort(key=lambda item: item[1], reverse=True)

    sections = {
        "A1": [["来源", "工单量"], ["热线工单", hotline_count], ["视频工单", video_count]],
        "D1": [
            ["服务方式", "解决率"],
            ["热线解决率", _percent_to_float(latest_result["热线解决率"])],
            ["视频解决率", _percent_to_float(latest_video_result["视频解决率"])],
        ],
        "G1": [["客户类型", "客户数量"], *customer_structure],
        "J1": [
            ["覆盖类型", "覆盖率"],
            ["整体客户", _percent_to_float(latest_result["热线（整体客户）覆盖率"])],
            ["双大客户", _percent_to_float(latest_result["热线（双大客户）覆盖率"])],
            ["县域+社办AB", _percent_to_float(latest_result["热线（县域+社办AB类客户）覆盖率"])],
            ["JC+社办CD", _percent_to_float(latest_result["热线（JC+社办）覆盖率"])],
        ],
    }

    for anchor, rows in sections.items():
        start_col = ord(re.match(r"[A-Z]+", anchor).group(0)) - ord("A") + 1
        start_row = int(re.search(r"\d+", anchor).group(0))
        for row_offset, row_values in enumerate(rows):
            for col_offset, value in enumerate(row_values):
                sheet.cell(start_row + row_offset, start_col + col_offset, value)

    for cell in ("E2", "E3", "K2", "K3", "K4", "K5"):
        sheet[cell].number_format = "0.00%"


def _add_dashboard_area(
    writer: pd.ExcelWriter,
    latest_result: dict[str, Any],
    latest_video_result: dict[str, Any],
    sheet_name: str = "计算结果",
) -> None:
    _write_dashboard_data_sheet(writer, latest_result, latest_video_result)
    worksheet = writer.book[sheet_name]
    data_sheet = writer.book["图表数据"]

    for row_index in range(12, 39):
        worksheet.row_dimensions[row_index].height = 24
    for column_index in range(1, 27):
        worksheet.column_dimensions[get_column_letter(column_index)].width = max(
            worksheet.column_dimensions[get_column_letter(column_index)].width or 0,
            13,
        )

    hotline_count = int(latest_result["热线工单数量"])
    video_count = int(latest_video_result["视频工单数量"])
    total_work_orders = hotline_count + video_count

    _style_dashboard_card(worksheet, "A12:E13", "热线解决率", str(latest_result["热线解决率"]), "1D4ED8")
    _style_dashboard_card(worksheet, "F12:J13", "视频解决率", str(latest_video_result["视频解决率"]), "047857")
    _style_dashboard_card(worksheet, "K12:O13", "热线整体覆盖率", str(latest_result["热线（整体客户）覆盖率"]), "0F766E")
    _style_dashboard_card(worksheet, "P12:T13", "总客户数", str(int(latest_result["总客户数量"])), "7C3AED")
    _style_dashboard_card(worksheet, "U12:Z13", "总工单量", str(total_work_orders), "EA580C")

    _add_chart_section_label(worksheet, "A14:H14", "【工单来源结构分析】", "展示工单来源构成。")
    _add_chart_section_label(worksheet, "J14:Q14", "【热线与视频解决率对比】", "展示不同服务方式的解决效果。")
    _add_chart_section_label(worksheet, "S14:Z14", "【客户结构分布分析】", "展示客户群体构成。")
    _add_chart_section_label(worksheet, "A28:M28", "【客户覆盖率分析】", "展示客户触达能力。")

    doughnut = DoughnutChart()
    doughnut.title = f"工单来源结构分析\n总工单量：{total_work_orders}"
    doughnut.holeSize = 55
    doughnut.add_data(Reference(data_sheet, min_col=2, min_row=1, max_row=3), titles_from_data=True)
    doughnut.set_categories(Reference(data_sheet, min_col=1, min_row=2, max_row=3))
    doughnut.dataLabels = DataLabelList()
    doughnut.dataLabels.showPercent = True
    doughnut.dataLabels.showLeaderLines = True
    doughnut.width = 14
    doughnut.height = 7
    worksheet.add_chart(doughnut, "A15")

    solve_chart = BarChart()
    solve_chart.type = "col"
    solve_chart.style = 10
    solve_chart.title = "热线与视频解决率对比"
    solve_chart.y_axis.title = "解决率"
    solve_chart.y_axis.numFmt = "0%"
    solve_chart.add_data(Reference(data_sheet, min_col=5, min_row=1, max_row=3), titles_from_data=True)
    solve_chart.set_categories(Reference(data_sheet, min_col=4, min_row=2, max_row=3))
    solve_chart.dataLabels = DataLabelList()
    solve_chart.dataLabels.showVal = True
    solve_chart.width = 14
    solve_chart.height = 7
    worksheet.add_chart(solve_chart, "J15")

    customer_chart = BarChart()
    customer_chart.type = "bar"
    customer_chart.style = 12
    customer_chart.title = "客户结构分布分析"
    customer_chart.x_axis.title = "客户数量"
    customer_chart.add_data(Reference(data_sheet, min_col=8, min_row=1, max_row=4), titles_from_data=True)
    customer_chart.set_categories(Reference(data_sheet, min_col=7, min_row=2, max_row=4))
    customer_chart.dataLabels = DataLabelList()
    customer_chart.dataLabels.showVal = True
    customer_chart.width = 14
    customer_chart.height = 7
    worksheet.add_chart(customer_chart, "S15")

    coverage_chart = BarChart()
    coverage_chart.type = "col"
    coverage_chart.style = 13
    coverage_chart.title = "热线客户覆盖率分析"
    coverage_chart.y_axis.title = "覆盖率"
    coverage_chart.y_axis.numFmt = "0%"
    coverage_chart.add_data(Reference(data_sheet, min_col=11, min_row=1, max_row=5), titles_from_data=True)
    coverage_chart.set_categories(Reference(data_sheet, min_col=10, min_row=2, max_row=5))
    coverage_chart.dataLabels = DataLabelList()
    coverage_chart.dataLabels.showVal = True
    coverage_chart.width = 24
    coverage_chart.height = 7
    worksheet.add_chart(coverage_chart, "A29")


def _json_safe_value(value: Any) -> Any:
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, pd.Timestamp):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        return float(value)
    return value


def _make_preview(df: pd.DataFrame, limit: int = 10) -> dict[str, Any]:
    preview_df = df.head(limit).copy()
    rows = []
    for record in preview_df.to_dict(orient="records"):
        rows.append({str(key): _json_safe_value(value) for key, value in record.items()})

    return {
        "columns": [str(column) for column in df.columns],
        "rows": rows,
        "limit": limit,
    }


def process_excels(work_order_file: UploadFile, quality_file: UploadFile) -> dict[str, Any]:
    work_df = _read_table(work_order_file, "工单报表")
    quality_df = _read_table(quality_file, "质量上升报表")

    product_line_col = _find_column(
        work_df, WORK_ORDER_ALIASES["product_line"], "工单报表.大产线"
    )
    status_col = _find_column(work_df, WORK_ORDER_ALIASES["status"], "工单报表.工单状态")
    work_ticket_col = _find_column(
        work_df, WORK_ORDER_ALIASES["ticket_id"], "工单报表.工单号"
    )
    quality_ticket_col = _find_column(
        quality_df, QUALITY_ALIASES["ticket_id"], "质量上升报表.前续工单"
    )

    product_line = work_df[product_line_col].map(_normalize_cell_text)
    status = work_df[status_col].map(_normalize_cell_text)

    pending_df = work_df[
        (product_line == TARGET_PRODUCT_LINE) & (status.isin(TARGET_STATUSES))
    ].copy()

    quality_ticket_ids: set[str] = set()
    for value in quality_df[quality_ticket_col]:
        quality_ticket_ids.update(_expand_ticket_ids(value))

    pending_ticket_ids = pending_df[work_ticket_col].map(_normalize_ticket_id)
    result_df = pending_df[~pending_ticket_ids.isin(quality_ticket_ids)].copy()

    job_id = uuid.uuid4().hex
    filename = f"工单报表-已筛选-{job_id}.xlsx"
    output_path = OUTPUT_DIR / filename

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        result_df.to_excel(writer, index=False, sheet_name=WORK_ORDER_SHEET_NAME)

    return {
        "job_id": job_id,
        "filename": filename,
        "download_url": f"/api/download/{job_id}",
        "stats": {
            "work_order_total": int(len(work_df)),
            "quality_total": int(len(quality_df)),
            "pending_total": int(len(pending_df)),
            "matched_quality_ticket_total": int(
                pending_ticket_ids.isin(quality_ticket_ids).sum()
            ),
            "result_total": int(len(result_df)),
        },
        "columns": {
            "product_line": str(product_line_col),
            "status": str(status_col),
            "work_ticket_id": str(work_ticket_col),
            "quality_ticket_id": str(quality_ticket_col),
        },
        "preview": _make_preview(result_df),
    }


def _assessment_provider_from_text(value: Any) -> str:
    text = _normalize_cell_text(value)
    if not text:
        return ""
    for provider, aliases in ASSESSMENT_PROVIDER_ALIASES.items():
        if any(alias in text for alias in aliases):
            return provider
    return ""


def _assessment_provider_from_staff(value: Any) -> str:
    normalized = _normalize_identifier(value)
    if not normalized:
        return ""
    normalized = re.sub(r"^CP", "", normalized, flags=re.IGNORECASE)
    normalized_lower = normalized.lower()
    return (
        ASSESSMENT_STAFF_PROVIDER_MAP.get(normalized_lower, "")
        or ASSESSMENT_STAFF_PROVIDER_MAP.get(f"s{normalized_lower}", "")
    )


def _duration_to_seconds(value: Any) -> float | None:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, (int, float, np.integer, np.floating)):
        if pd.isna(value):
            return None
        return float(value)
    if hasattr(value, "hour") and hasattr(value, "minute") and hasattr(value, "second"):
        return float(value.hour * 3600 + value.minute * 60 + value.second)

    text = str(value).strip()
    if not text:
        return None
    if text.endswith("秒"):
        text = text[:-1].strip()
        try:
            return float(text)
        except ValueError:
            return None
    if text.endswith("min"):
        try:
            return float(text[:-3].strip()) * 60
        except ValueError:
            return None
    if ":" in text:
        try:
            parts = [float(part) for part in text.split(":")]
        except ValueError:
            return None
        if len(parts) == 3:
            return parts[0] * 3600 + parts[1] * 60 + parts[2]
        if len(parts) == 2:
            return parts[0] * 60 + parts[1]
    try:
        return float(text)
    except ValueError:
        return None


def _datetime_diff_seconds(end_value: Any, start_value: Any) -> float | None:
    end = pd.to_datetime(end_value, errors="coerce")
    start = pd.to_datetime(start_value, errors="coerce")
    if pd.isna(end) or pd.isna(start):
        return None
    return float((end - start).total_seconds())


def _format_duration_hours(seconds: float | None) -> str:
    if seconds is None:
        return "N/A"
    return f"{seconds / 3600:.2f}小时"


def _assessment_ratio(numerator: int | float, denominator: int | float) -> str:
    return _format_percent(_safe_divide(numerator, denominator))


def _build_assessment_branch_result(msp_df: pd.DataFrame) -> pd.DataFrame:
    grouped = []
    status = msp_df["工单状态"].map(_normalize_cell_text)
    branch = msp_df["分公司"].map(_normalize_cell_text)
    for branch_name in sorted(value for value in branch.unique() if value):
        branch_mask = branch == branch_name
        total = int(branch_mask.sum())
        canceled = int((branch_mask & (status == "已取消")).sum())
        grouped.append({
            "分公司": branch_name,
            "工单总数量": total,
            "已取消工单数量": canceled,
            "分公司工单取消率": _assessment_ratio(canceled, total),
        })
    return pd.DataFrame(grouped, columns=["分公司", "工单总数量", "已取消工单数量", "分公司工单取消率"])


def _build_assessment_overall_result(msp_df: pd.DataFrame) -> pd.DataFrame:
    channel = msp_df["申告渠道"].map(_normalize_cell_text)
    status = msp_df["工单状态"].map(_normalize_cell_text)
    online_mask = channel.isin(ASSESSMENT_ONLINE_CHANNELS)
    non_cancel_online_mask = online_mask & (status != "已取消")

    timely_denominator = int(non_cancel_online_mask.sum())
    timely_numerator = 0
    repair_seconds_total = 0.0
    repair_denominator = int(online_mask.sum())
    for _, row in msp_df[online_mask].iterrows():
        if _normalize_cell_text(row["工单状态"]) != "已取消":
            appointment_seconds = _datetime_diff_seconds(row["预约完成时间"], row["首次派单时间"])
            if appointment_seconds is not None and appointment_seconds <= 1800:
                timely_numerator += 1
        repair_seconds = _datetime_diff_seconds(row["服务结束时间"], row["首次派单时间"])
        if repair_seconds is not None and repair_seconds > 0:
            repair_seconds_total += repair_seconds

    average_repair_seconds = _safe_divide(repair_seconds_total, repair_denominator)
    return pd.DataFrame([
        {
            "指标": "在线服务工单预约及时率",
            "分子说明": "预约完成时间-首次派单时间≤30分钟的在线服务工单数（已取消不计）",
            "分子值": timely_numerator,
            "分母说明": "在线服务渠道工单数（已取消不计）",
            "分母值": timely_denominator,
            "结果": _assessment_ratio(timely_numerator, timely_denominator),
        },
        {
            "指标": "平均修复时长（线上）",
            "分子说明": "在线服务渠道工单服务结束时间-首次派单时间的总时长",
            "分子值": _format_duration_hours(repair_seconds_total),
            "分母说明": "在线服务渠道工单数量",
            "分母值": repair_denominator,
            "结果": _format_duration_hours(average_repair_seconds),
        },
    ])


def _calculate_video_answer_rate(provider_df: pd.DataFrame) -> tuple[int, int, str]:
    answered = provider_df["是否接通"].map(_normalize_cell_text)
    ring_seconds = pd.to_numeric(provider_df["振铃时长"].map(_duration_to_seconds), errors="coerce")
    fail_reason = provider_df["失败原因"].map(_normalize_cell_text)
    numerator = int(((answered == "是") & (ring_seconds <= 15)).sum())
    b_count = int(((answered == "是") & (ring_seconds > 15)).sum())
    c_count = int(((answered == "否") & (ring_seconds >= 30)).sum())
    d_count = int(((answered == "否") & (fail_reason == "canceled") & (ring_seconds >= 15)).sum())
    denominator = numerator + b_count + c_count + d_count
    return numerator, denominator, _assessment_ratio(numerator, denominator)


def _calculate_hotline_satisfaction(provider_df: pd.DataFrame) -> tuple[int, str]:
    scores = pd.to_numeric(provider_df["短信满意度结果"], errors="coerce")
    a_count = int((scores == 1).sum())
    b_count = int((scores == 2).sum())
    c_count = int((scores == 3).sum())
    denominator = a_count + b_count + c_count
    score = _safe_divide(a_count * 100 + b_count * 80 + c_count * 60, denominator * 100)
    return denominator, _format_percent(score)


def _calculate_video_satisfaction(provider_df: pd.DataFrame) -> tuple[int, str]:
    scores = pd.to_numeric(provider_df["评价"], errors="coerce")
    valid = scores[scores.isin([1, 2, 3, 4, 5])]
    denominator = int(len(valid))
    score = _safe_divide(float(valid.sum()), denominator * 5)
    return denominator, _format_percent(score)


def _build_assessment_provider_result(
    mcc_call_df: pd.DataFrame,
    video_service_df: pd.DataFrame,
    mcc_ticket_df: pd.DataFrame,
    crm_video_df: pd.DataFrame,
    quality_df: pd.DataFrame,
) -> pd.DataFrame:
    mcc_call = mcc_call_df.copy()
    mcc_call["__provider"] = mcc_call["相关客服"].map(_assessment_provider_from_text)

    video_service = video_service_df.copy()
    video_service["__provider"] = video_service["客服工号"].map(_assessment_provider_from_staff)

    mcc_ticket = mcc_ticket_df.copy()
    mcc_ticket["__provider"] = mcc_ticket["受理客服"].map(_assessment_provider_from_text)

    crm_video = crm_video_df.copy()
    crm_video["__provider"] = crm_video.apply(
        lambda row: (
            _assessment_provider_from_text(row["负责GTS编号"])
            or _assessment_provider_from_staff(row["负责GTS编号"])
            or _assessment_provider_from_text(row.get("负责GTS", ""))
            or _assessment_provider_from_staff(row.get("负责GTS", ""))
        ),
        axis=1,
    )

    quality = quality_df.copy()
    quality["__provider"] = quality["代理商"].map(_assessment_provider_from_text)
    quality_deduction_col = _find_required_column_by_prefix(quality, "扣分", "每日质检记录表")

    rows = []
    for provider in ASSESSMENT_PROVIDERS:
        provider_calls = mcc_call[mcc_call["__provider"] == provider]
        hotline_ring_seconds = pd.to_numeric(provider_calls["响铃时间"].map(_duration_to_seconds), errors="coerce")
        hotline_answer_denominator = int(len(provider_calls))
        hotline_answer_numerator = int((hotline_ring_seconds <= 15).sum())

        provider_video_service = video_service[video_service["__provider"] == provider]
        video_answer_numerator, video_answer_denominator, video_answer_rate = _calculate_video_answer_rate(
            provider_video_service
        )

        provider_mcc_tickets = mcc_ticket[mcc_ticket["__provider"] == provider]
        hotline_duration_total = 0.0
        for _, row in provider_mcc_tickets.iterrows():
            duration = _datetime_diff_seconds(row["关闭时间"], row["创建时间"])
            if duration is not None and duration > 0:
                hotline_duration_total += duration

        video_duration_total = float(
            provider_video_service["通话时间"].map(_duration_to_seconds).fillna(0).sum()
        )
        processing_denominator = int(len(provider_mcc_tickets) + len(provider_video_service))
        average_processing_seconds = _safe_divide(
            hotline_duration_total + video_duration_total,
            processing_denominator,
        )

        hotline_satisfaction_count, hotline_satisfaction = _calculate_hotline_satisfaction(provider_mcc_tickets)
        video_satisfaction_count, video_satisfaction = _calculate_video_satisfaction(provider_video_service)

        provider_quality = quality[quality["__provider"] == provider]
        quality_total = int(len(provider_quality))
        quality_failed = int(provider_quality[quality_deduction_col].map(_normalize_cell_text).ne("").sum())
        quality_passed = quality_total - quality_failed

        mcc_status = provider_mcc_tickets["受理单(service call)状态"].map(_normalize_cell_text)
        hotline_solved = int((mcc_status == "技术支持解决").sum())
        hotline_solution_denominator = int(mcc_status.isin({"技术支持解决", "派工完成"}).sum())

        provider_crm_video = crm_video[crm_video["__provider"] == provider]
        video_status = provider_crm_video["用户状态"].map(_normalize_cell_text)
        video_solved = int((video_status == "技术支持解决").sum())
        video_solution_denominator = int(video_status.isin({"技术支持解决", "转FSM平台派工"}).sum())

        queue_seconds = pd.to_numeric(provider_video_service["排队时长"].map(_duration_to_seconds), errors="coerce")
        video_busy_numerator = int((queue_seconds > 3).sum())
        video_busy_denominator = int(len(provider_video_service))

        rows.append({
            "分包商": provider,
            "热线15秒接起率": _assessment_ratio(hotline_answer_numerator, hotline_answer_denominator),
            "热线15秒分子": hotline_answer_numerator,
            "热线15秒分母": hotline_answer_denominator,
            "视频15秒接起率": video_answer_rate,
            "视频15秒分子": video_answer_numerator,
            "视频15秒分母": video_answer_denominator,
            "平均在线工单处理时长": _format_duration_hours(average_processing_seconds),
            "在线处理工单数": processing_denominator,
            "热线满意度": hotline_satisfaction,
            "热线满意度样本数": hotline_satisfaction_count,
            "视频满意度": video_satisfaction,
            "视频满意度样本数": video_satisfaction_count,
            "质检合格率": _assessment_ratio(quality_passed, quality_total),
            "质检合格数": quality_passed,
            "质检总数": quality_total,
            "热线解决率": _assessment_ratio(hotline_solved, hotline_solution_denominator),
            "热线解决分子": hotline_solved,
            "热线解决分母": hotline_solution_denominator,
            "视频解决率": _assessment_ratio(video_solved, video_solution_denominator),
            "视频解决分子": video_solved,
            "视频解决分母": video_solution_denominator,
            "视频占线率": _assessment_ratio(video_busy_numerator, video_busy_denominator),
            "视频占线分子": video_busy_numerator,
            "视频占线分母": video_busy_denominator,
        })

    return pd.DataFrame(rows)


def _style_assessment_workbook(writer: pd.ExcelWriter) -> None:
    workbook = writer.book
    header_fill = PatternFill("solid", fgColor="EAF3FF")
    title_fill = PatternFill("solid", fgColor="DFF7EF")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for sheet_name in ("计算结果", "计算逻辑"):
        worksheet = workbook[sheet_name]
        for row in worksheet.iter_rows():
            first_value = row[0].value
            is_title_row = isinstance(first_value, str) and first_value in {"分公司指标", "整体指标", "分包商指标"}
            previous_first_value = worksheet.cell(row=max(row[0].row - 1, 1), column=1).value
            is_header_row = row[0].row == 1 or previous_first_value in {"分公司指标", "整体指标", "分包商指标"}
            for cell in row:
                if cell.value is None:
                    continue
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
                if is_title_row:
                    cell.font = Font(bold=True, size=14, color="064E3B")
                    cell.fill = title_fill
                elif is_header_row or cell.row == 1:
                    cell.font = Font(bold=True, size=11, color="0F172A")
                    cell.fill = header_fill
        for column_index in range(1, min(worksheet.max_column, 30) + 1):
            worksheet.column_dimensions[get_column_letter(column_index)].width = 18

    result_ws = workbook["计算结果"]
    for row in result_ws.iter_rows():
        first_value = row[0].value
        if isinstance(first_value, str) and first_value in {"分公司指标", "整体指标", "分包商指标"}:
            for cell in row:
                cell.font = Font(bold=True, size=14, color="064E3B")
                cell.fill = title_fill
    result_ws.freeze_panes = "A3"


def _write_assessment_logic_sheet(writer: pd.ExcelWriter) -> None:
    rows = [
        [1, "分公司工单取消率", "按分公司统计 MSP 工单总数和工单状态为“已取消”的数量，取消数量/工单总数。", "MSP工单总表"],
        [2, "在线服务工单预约及时率", "筛选热线申告、微信视频申告、一键报修申告、微信申告、远程申告且未取消的 MSP 工单；预约完成时间-首次派单时间≤30分钟为及时。", "MSP工单总表"],
        [3, "在线工单15秒接起率", "热线按相关客服匹配分包商并统计响铃时间≤15秒；视频按客服工号匹配分包商，按接通/未接通及失败原因组成分母并统计15秒内接通。", "MCC通话记录；视频服务记录"],
        [4, "平均在线工单处理时长", "热线取关闭时间-创建时间，视频取通话时间；按分包商汇总总时长后除以热线+视频工单数量。", "MCC热线工单；视频服务记录"],
        [5, "在线工单满意度", "热线按短信满意度结果1/2/3折算100/80/60分；视频按评价1-5分折算为满分5分占比。", "MCC热线工单；视频服务记录"],
        [6, "质检合格率", "按分包商统计质检记录；扣分字段为空为合格，合格数/质检总数。", "每日质检记录表"],
        [7, "平均修复时长（线上）", "筛选在线服务渠道 MSP 工单，服务结束时间-首次派单时间的总时长除以工单数量。", "MSP工单总表"],
        [8, "热线解决率", "按分包商筛选 MCC 热线工单，技术支持解决数量/（技术支持解决+派工完成）。", "MCC热线工单"],
        [9, "视频解决率", "按负责GTS编号匹配分包商，技术支持解决数量/（技术支持解决+转FSM平台派工）。", "CRM视频工单"],
        [10, "视频占线率", "按客服工号匹配分包商，排队时长>3秒的视频记录数/该分包商视频记录总数。", "视频服务记录"],
    ]
    logic_df = pd.DataFrame(rows, columns=["序号", "指标", "计算逻辑", "表格来源"])
    logic_df.to_excel(writer, index=False, sheet_name="计算逻辑")


def _write_assessment_results(
    writer: pd.ExcelWriter,
    branch_df: pd.DataFrame,
    overall_df: pd.DataFrame,
    provider_df: pd.DataFrame,
) -> None:
    sheet_name = "计算结果"
    startrow = 1
    branch_df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=startrow)
    worksheet = writer.sheets[sheet_name]
    worksheet.cell(row=1, column=1, value="分公司指标")
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(branch_df.columns))

    overall_title_row = startrow + len(branch_df) + 4
    worksheet.cell(row=overall_title_row, column=1, value="整体指标")
    worksheet.merge_cells(start_row=overall_title_row, start_column=1, end_row=overall_title_row, end_column=len(overall_df.columns))
    overall_df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=overall_title_row)

    provider_title_row = overall_title_row + len(overall_df) + 4
    worksheet.cell(row=provider_title_row, column=1, value="分包商指标")
    worksheet.merge_cells(start_row=provider_title_row, start_column=1, end_row=provider_title_row, end_column=len(provider_df.columns))
    provider_df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=provider_title_row)


def process_online_assessment_excels(
    msp_file: UploadFile,
    mcc_call_file: UploadFile,
    video_service_file: UploadFile,
    mcc_ticket_file: UploadFile,
    crm_video_file: UploadFile,
    quality_file: UploadFile,
    include_source_sheets: bool = False,
) -> dict[str, Any]:
    msp_df = _read_assessment_table(msp_file, "MSP工单总表", "MSP工单原始数据")
    mcc_call_df = _read_assessment_table(mcc_call_file, "MCC通话记录")
    video_service_df = _read_assessment_table(video_service_file, "视频服务记录")
    mcc_ticket_df = _read_assessment_table(mcc_ticket_file, "MCC热线工单")
    crm_video_df = _read_assessment_table(crm_video_file, "CRM视频工单")
    quality_df = _read_assessment_table(quality_file, "每日质检记录表", "质检表格")

    _ensure_columns(msp_df, ["工单状态", "分公司", "申告渠道", "首次派单时间", "预约完成时间", "服务结束时间"], "MSP工单总表")
    _ensure_columns(mcc_call_df, ["相关客服", "响铃时间"], "MCC通话记录")
    _ensure_columns(video_service_df, ["客服工号", "振铃时长", "是否接通", "失败原因", "通话时间", "评价", "排队时长"], "视频服务记录")
    _ensure_columns(mcc_ticket_df, ["受理客服", "创建时间", "关闭时间", "短信满意度结果", "受理单(service call)状态"], "MCC热线工单")
    _ensure_columns(crm_video_df, ["负责GTS编号", "负责GTS", "用户状态"], "CRM视频工单")
    _ensure_columns(quality_df, ["代理商"], "每日质检记录表")
    _find_required_column_by_prefix(quality_df, "扣分", "每日质检记录表")

    branch_df = _build_assessment_branch_result(msp_df)
    overall_df = _build_assessment_overall_result(msp_df)
    provider_df = _build_assessment_provider_result(
        mcc_call_df=mcc_call_df,
        video_service_df=video_service_df,
        mcc_ticket_df=mcc_ticket_df,
        crm_video_df=crm_video_df,
        quality_df=quality_df,
    )

    job_id = uuid.uuid4().hex
    filename = f"在线服务考核指标计算表_{job_id}.xlsx"
    output_path = OUTPUT_DIR / filename
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        if include_source_sheets:
            msp_df.to_excel(writer, index=False, sheet_name="源表-MSP工单总表")
            mcc_call_df.to_excel(writer, index=False, sheet_name="源表-MCC通话记录")
            video_service_df.to_excel(writer, index=False, sheet_name="源表-视频服务记录")
            mcc_ticket_df.to_excel(writer, index=False, sheet_name="源表-MCC热线工单")
            crm_video_df.to_excel(writer, index=False, sheet_name="源表-CRM视频工单")
            quality_df.to_excel(writer, index=False, sheet_name="源表-每日质检记录")
        _write_assessment_results(writer, branch_df, overall_df, provider_df)
        _write_assessment_logic_sheet(writer)
        _style_assessment_workbook(writer)

    return {
        "job_id": job_id,
        "filename": filename,
        "download_url": f"/api/online-assessment/download/{job_id}",
        "stats": {
            "msp_total": int(len(msp_df)),
            "mcc_call_total": int(len(mcc_call_df)),
            "video_service_total": int(len(video_service_df)),
            "mcc_ticket_total": int(len(mcc_ticket_df)),
            "crm_video_total": int(len(crm_video_df)),
            "quality_total": int(len(quality_df)),
            "include_source_sheets": include_source_sheets,
        },
        "preview": _make_preview(provider_df),
    }


def process_online_business_excels(
    mcc_file: UploadFile,
    video_file: UploadFile,
    msp_file: UploadFile,
    ivd_customer_file: UploadFile,
    include_source_sheets: bool = False,
) -> dict[str, Any]:
    _validate_xlsx_upload_any(
        mcc_file,
        "MCC热线原始数据",
        ["MCC热线原始数据", "MCC热线工单", "MCC热线"],
    )
    _validate_xlsx_upload_any(
        video_file,
        "视频工单原始数据",
        ["视频工单原始数据", "CRM视频工单", "视频工单"],
    )
    _validate_xlsx_upload_any(
        msp_file,
        "MSP工单原始数据",
        ["MSP工单原始数据", "MSP工单总表", "MSP工单"],
    )
    _validate_xlsx_upload_any(
        ivd_customer_file,
        "IVD客户群表",
        ["IVD客户群表", "IVD客户群数量2025"],
    )

    mcc_df = _read_required_excel_sheet(mcc_file, "MCC热线原始数据", ONLINE_BUSINESS_SHEETS["mcc"])
    video_df = _read_required_excel_sheet(video_file, "视频工单原始数据", ONLINE_BUSINESS_SHEETS["video"])
    msp_df = _read_required_excel_sheet(msp_file, "MSP工单原始数据", ONLINE_BUSINESS_SHEETS["msp"])
    ivd_sheets = _read_ivd_customer_sheets(ivd_customer_file)

    _ensure_columns(mcc_df, ONLINE_REQUIRED_COLUMNS["mcc"], "MCC热线原始数据")
    _ensure_columns(video_df, ONLINE_REQUIRED_COLUMNS["video"], "视频工单原始数据")
    _ensure_columns(msp_df, ONLINE_REQUIRED_COLUMNS["msp"], "MSP工单原始数据")
    _ensure_columns(
        ivd_sheets[ONLINE_BUSINESS_SHEETS["ivd_double"]],
        ONLINE_REQUIRED_COLUMNS["ivd_double"],
        "IVD客户群数量2025.双大",
    )
    _ensure_columns(
        ivd_sheets[ONLINE_BUSINESS_SHEETS["ivd_ab"]],
        ONLINE_REQUIRED_COLUMNS["ivd_ab"],
        "IVD客户群数量2025.县域+社办AB类",
    )
    _ensure_columns(
        ivd_sheets[ONLINE_BUSINESS_SHEETS["ivd_jc"]],
        ONLINE_REQUIRED_COLUMNS["ivd_jc"],
        "IVD客户群数量2025.JC+社办CD类",
    )

    years = _detect_online_business_years(mcc_df, video_df, msp_df)
    target_year = max(years)
    mcc_year_df = _filter_year(mcc_df, "创建时间", target_year)

    double_ids, ab_ids = _build_customer_level_lookup(ivd_sheets)
    _, ab_names = _build_customer_name_lookup(ivd_sheets)
    dedup_detail_df = _build_mcc_customer_dedup_detail(mcc_year_df, double_ids, ab_ids)

    double_customer_count = int(len(ivd_sheets[ONLINE_BUSINESS_SHEETS["ivd_double"]]))
    ab_customer_count = int(len(ivd_sheets[ONLINE_BUSINESS_SHEETS["ivd_ab"]]))
    jc_customer_total = int(
        pd.to_numeric(
            ivd_sheets[ONLINE_BUSINESS_SHEETS["ivd_jc"]]["客户数量"], errors="coerce"
        ).fillna(0).sum()
    )
    ivd_customer_total = double_customer_count + ab_customer_count + jc_customer_total

    result_df = _build_online_business_result_table(
        years=years,
        mcc_df=mcc_df,
        video_df=video_df,
        msp_df=msp_df,
        ivd_sheets=ivd_sheets,
        double_ids=double_ids,
        ab_ids=ab_ids,
    )
    video_result_df = _build_video_business_result_table(
        years=years,
        mcc_df=mcc_df,
        video_df=video_df,
        msp_df=msp_df,
        ivd_sheets=ivd_sheets,
        double_ids=double_ids,
        ab_ids=ab_ids,
        ab_names=ab_names,
    )
    combined_coverage_result_df = _build_combined_coverage_result_table(
        years=years,
        mcc_df=mcc_df,
        video_df=video_df,
        msp_df=msp_df,
        ivd_sheets=ivd_sheets,
        double_ids=double_ids,
        ab_ids=ab_ids,
        ab_names=ab_names,
    )
    latest_result = result_df.iloc[-1].to_dict()
    latest_video_result = video_result_df.iloc[-1].to_dict()
    latest_combined_result = combined_coverage_result_df.iloc[-1].to_dict()

    job_id = uuid.uuid4().hex
    filename = f"在线服务项目目标计算表_{target_year}_{job_id}.xlsx"
    output_path = OUTPUT_DIR / filename

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        if include_source_sheets:
            mcc_df.to_excel(writer, index=False, sheet_name="MCC热线原始数据")
            video_df.to_excel(writer, index=False, sheet_name="视频工单原始数据")
            msp_df.to_excel(writer, index=False, sheet_name="MSP工单原始数据")
            _write_ivd_customer_source_sheets(writer, ivd_sheets)
            dedup_detail_df.to_excel(writer, index=False, sheet_name="MCC热线客户去重明细")
        hotline_title_row = 1
        hotline_header_row = 2
        hotline_startrow = hotline_header_row - 1
        result_df.to_excel(
            writer,
            index=False,
            sheet_name="计算结果",
            startrow=hotline_startrow,
        )
        video_title_row = hotline_header_row + len(result_df) + 2
        video_header_row = video_title_row + 1
        video_startrow = video_header_row - 1
        video_result_df.to_excel(
            writer,
            index=False,
            sheet_name="计算结果",
            startrow=video_startrow,
        )
        combined_title_row = video_header_row + len(video_result_df) + 2
        combined_header_row = combined_title_row + 1
        combined_startrow = combined_header_row - 1
        combined_coverage_result_df.to_excel(
            writer,
            index=False,
            sheet_name="计算结果",
            startrow=combined_startrow,
        )
        _add_result_table_titles(
            writer,
            table_specs=[
                ("热线数据", hotline_title_row, len(result_df.columns)),
                ("视频数据", video_title_row, len(video_result_df.columns)),
                ("热线+视频数据", combined_title_row, len(combined_coverage_result_df.columns)),
            ],
        )
        _apply_online_result_formulas(writer, header_row=hotline_header_row, row_count=len(result_df))
        _apply_video_result_formulas(
            writer,
            header_row=video_header_row,
            row_count=len(video_result_df),
        )
        _apply_combined_coverage_result_formulas(
            writer,
            header_row=combined_header_row,
            row_count=len(combined_coverage_result_df),
        )
        _style_online_result_sheet(
            writer,
            header_rows=[hotline_header_row, video_header_row, combined_header_row],
            title_rows=[hotline_title_row, video_title_row, combined_title_row],
        )
        _add_calculation_logic_sheet(writer)
        _add_dashboard_area(
            writer,
            latest_result=latest_result,
            latest_video_result=latest_video_result,
        )

    return {
        "job_id": job_id,
        "filename": filename,
        "download_url": f"/api/online-business/download/{job_id}",
        "target_year": target_year,
        "stats": {
            "mcc_total": int(len(mcc_df)),
            "video_total": int(len(video_df)),
            "msp_total": int(len(msp_df)),
            "result_years": years,
            "latest_year": target_year,
            "ivd_customer_total": ivd_customer_total,
            "latest_year_coverage_customer_total": int(latest_result["热线覆盖客户数量"]),
            "latest_year_jc_customer_total": int(latest_result["热线覆盖（JC+社办）客户数量"]),
            "latest_year_video_coverage_customer_total": int(latest_video_result["视频覆盖客户数量"]),
            "latest_year_video_jc_customer_total": int(latest_video_result["视频覆盖（JC+社办）客户数量"]),
            "latest_year_combined_coverage_customer_total": int(latest_combined_result["视频+热线覆盖客户数量"]),
            "latest_year_combined_jc_customer_total": int(latest_combined_result["视频+热线覆盖（JC+社办）客户数量"]),
            "dedup_customer_total": int(len(dedup_detail_df)),
            "include_source_sheets": include_source_sheets,
        },
        "preview": _make_preview(result_df),
    }


@app.get("/api/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/api/license/machine-code")
def license_machine_code() -> dict[str, Any]:
    return {"machine_code": get_machine_code(), "enabled": is_license_required()}


@app.get("/api/license/status")
def license_status() -> dict[str, Any]:
    info = license_to_info(current_license())
    if info.get("enabled") and info.get("status") == "active":
        try:
            return require_valid_license()
        except LicenseError as exc:
            blocked = dict(info)
            blocked.update({"status": "invalid", "message": exc.message, "license_required": True})
            return blocked
    return info


@app.post("/api/license/activate")
def license_activate(payload: dict[str, Any]) -> dict[str, Any]:
    code = str(payload.get("license_code", "")).strip()
    if not code:
        raise HTTPException(status_code=400, detail="请输入注册码")
    try:
        info = activate_license(code)
    except LicenseError as exc:
        raise HTTPException(status_code=exc.status_code, detail=exc.message) from exc
    return {"success": True, "license": info}


@app.post("/api/license/verify")
def license_verify() -> dict[str, Any]:
    try:
        info = require_valid_license()
    except LicenseError as exc:
        raise HTTPException(status_code=exc.status_code, detail=exc.message) from exc
    return {"success": True, "license": info}


@app.post("/api/process")
def process_files(
    work_order_file: UploadFile = File(...),
    quality_file: UploadFile = File(...),
) -> dict[str, Any]:
    _require_feature(FEATURES["TIMEOUT_FILTER"])
    return process_excels(work_order_file, quality_file)


@app.post("/api/online-business/process")
def process_online_business_files(
    mcc_file: UploadFile = File(...),
    video_file: UploadFile = File(...),
    msp_file: UploadFile = File(...),
    ivd_customer_file: UploadFile = File(...),
    include_source_sheets: bool = Form(False),
) -> dict[str, Any]:
    _require_feature(FEATURES["ONLINE_SERVICE_TARGET"])
    return process_online_business_excels(
        mcc_file=mcc_file,
        video_file=video_file,
        msp_file=msp_file,
        ivd_customer_file=ivd_customer_file,
        include_source_sheets=include_source_sheets,
    )


@app.post("/api/online-assessment/process")
def process_online_assessment_files(
    msp_file: UploadFile = File(...),
    mcc_call_file: UploadFile = File(...),
    video_service_file: UploadFile = File(...),
    mcc_ticket_file: UploadFile = File(...),
    crm_video_file: UploadFile = File(...),
    quality_file: UploadFile = File(...),
    include_source_sheets: bool = Form(False),
) -> dict[str, Any]:
    _require_feature(FEATURES["ONLINE_SERVICE_KPI"])
    return process_online_assessment_excels(
        msp_file=msp_file,
        mcc_call_file=mcc_call_file,
        video_service_file=video_service_file,
        mcc_ticket_file=mcc_ticket_file,
        crm_video_file=crm_video_file,
        quality_file=quality_file,
        include_source_sheets=include_source_sheets,
    )


@app.get("/api/download/{job_id}")
def download_file(job_id: str) -> FileResponse:
    _require_feature(FEATURES["TIMEOUT_FILTER"])
    _require_feature(FEATURES["EXPORT_EXCEL"])
    if not re.fullmatch(r"[a-f0-9]{32}", job_id):
        raise HTTPException(status_code=400, detail="无效的下载编号")

    matches = list(OUTPUT_DIR.glob(f"工单报表-已筛选-{job_id}.xlsx"))
    if not matches:
        raise HTTPException(status_code=404, detail="导出文件不存在或已被清理")

    return FileResponse(
        matches[0],
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="工单报表-已筛选.xlsx",
    )


@app.get("/api/online-business/download/{job_id}")
def download_online_business_file(job_id: str) -> FileResponse:
    _require_feature(FEATURES["ONLINE_SERVICE_TARGET"])
    _require_feature(FEATURES["EXPORT_EXCEL"])
    if not re.fullmatch(r"[a-f0-9]{32}", job_id):
        raise HTTPException(status_code=400, detail="无效的下载编号")

    matches = list(OUTPUT_DIR.glob(f"在线服务项目目标计算表_*_{job_id}.xlsx"))
    if not matches:
        raise HTTPException(status_code=404, detail="在线服务项目目标计算表不存在或已被清理")

    year_match = re.search(r"在线服务项目目标计算表_(\d{4})_", matches[0].name)
    year = year_match.group(1) if year_match else "年度"
    return FileResponse(
        matches[0],
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"在线服务项目目标计算表_{year}.xlsx",
    )


@app.get("/api/online-assessment/download/{job_id}")
def download_online_assessment_file(job_id: str) -> FileResponse:
    _require_feature(FEATURES["ONLINE_SERVICE_KPI"])
    _require_feature(FEATURES["EXPORT_EXCEL"])
    if not re.fullmatch(r"[a-f0-9]{32}", job_id):
        raise HTTPException(status_code=400, detail="无效的下载编号")

    matches = list(OUTPUT_DIR.glob(f"在线服务考核指标计算表_{job_id}.xlsx"))
    if not matches:
        raise HTTPException(status_code=404, detail="在线服务考核指标计算表不存在或已被清理")

    return FileResponse(
        matches[0],
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="在线服务考核指标计算表.xlsx",
    )


@app.get("/", include_in_schema=False)
def serve_frontend():
    index_file = FRONTEND_DIST_DIR / "index.html"
    if index_file.exists():
        return FileResponse(index_file)

    return HTMLResponse(
        """
        <!doctype html>
        <html lang="zh-CN">
          <head>
            <meta charset="utf-8" />
            <title>技术支持效率平台</title>
            <style>
              body {
                margin: 0;
                min-height: 100vh;
                display: grid;
                place-items: center;
                background: #06111f;
                color: #e8f3ff;
                font-family: -apple-system, BlinkMacSystemFont, "PingFang SC",
                  "Microsoft YaHei", sans-serif;
              }
              main {
                width: min(680px, calc(100% - 32px));
                padding: 28px;
                border: 1px solid rgba(112, 198, 255, 0.28);
                border-radius: 8px;
                background: rgba(8, 25, 43, 0.9);
              }
              code {
                color: #65e7ff;
              }
            </style>
          </head>
          <body>
            <main>
              <h1>技术支持效率平台后端服务已启动</h1>
              <p>如果要打开完整前端页面，请在项目根目录执行：</p>
              <p><code>cd frontend && npm run dev</code></p>
              <p>然后访问：<code>http://127.0.0.1:5173</code></p>
              <p>或执行 <code>cd frontend && npm run build</code> 后刷新当前后端地址。</p>
            </main>
          </body>
        </html>
        """
    )


@app.get("/{full_path:path}", include_in_schema=False)
def serve_frontend_routes(full_path: str):
    if full_path.startswith("api/"):
        raise HTTPException(status_code=404, detail="接口不存在")

    index_file = FRONTEND_DIST_DIR / "index.html"
    if not index_file.exists():
        raise HTTPException(status_code=404, detail="前端页面未构建")

    return FileResponse(index_file)


if __name__ == "__main__":
    import uvicorn

    _prepare_runtime_streams()
    port = _find_free_port(8000)
    url = f"http://127.0.0.1:{port}"
    print(f"{APP_NAME} 已启动：{url}", flush=True)
    threading.Timer(1.2, lambda: webbrowser.open(url)).start()
    uvicorn.run(
        app,
        host="127.0.0.1",
        port=port,
        http="h11",
        log_config=None,
        access_log=False,
    )
